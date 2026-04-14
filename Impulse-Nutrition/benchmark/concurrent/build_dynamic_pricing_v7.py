"""
Build v7 — single-tab dynamic pricing for Impulse whey gamme.

One tab "Pricing Dynamique" with 3 stacked tables :
  1. Whey Concentrate (10 options, hardcoded prices)
  2. Whey Isolate    (10 options, dynamic from Concentrate retenu)
  3. Whey Recovery   (10 options, dynamic from Concentrate + Isolate retenus)

Each table has a dropdown selection (1-10), a "Prix retenu" cell that uses
INDEX/MATCH to retrieve the selected option. Changing the Concentrate
selection auto-cascades Isolate+Recovery via Excel formulas.

Zero hardcoded parameters for Isolate/Recovery — all computed from PR,
remise, market benchmarks, and cascade chain.

Design : Bloomberg-meets-Notion. Sharp dividers, mono numbers, colored
section headers per product (blue Concentrate / emerald Isolate / red
Recovery), amber-highlighted selection cells, zebra rows.
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path("/Users/antoinechabrat/Downloads/Copie de benchmark_whey_concurrent 5 1.xlsx")
DST = Path(
    "/Users/antoinechabrat/Documents/SmallProject/Impulse-Nutrition/benchmark/"
    "concurrent/benchmark_whey_concurrent_v7_dynamic.xlsx"
)

# ─── Design tokens ─────────────────────────────────────────────────────────
COLOR_INK        = "0F172A"   # slate-900
COLOR_MUTED      = "64748B"   # slate-500
COLOR_BORDER     = "CBD5E1"   # slate-300
COLOR_LIGHT      = "F1F5F9"   # slate-100
COLOR_WHITE      = "FFFFFF"
COLOR_ACCENT_RED = "DC2626"   # red-600 — for selection highlight
COLOR_AMBER      = "FEF3C7"   # amber-100 — for selected rows
COLOR_AMBER_STRONG = "F59E0B" # amber-500

# Product colors
COLOR_CONCENTRATE = "2563EB"  # blue-600
COLOR_ISOLATE     = "059669"  # emerald-600
COLOR_RECOVERY    = "B91C1C"  # red-700

FONT_TITLE    = Font(name="Arial", size=18, bold=True, color=COLOR_INK)
FONT_SUBTITLE = Font(name="Arial", size=10, italic=True, color=COLOR_MUTED)
FONT_SECTION  = Font(name="Arial", size=13, bold=True, color=COLOR_WHITE)
FONT_LABEL    = Font(name="Arial", size=10, bold=True, color=COLOR_INK)
FONT_LABEL_MUTED = Font(name="Arial", size=9, italic=True, color=COLOR_MUTED)
FONT_VALUE    = Font(name="Arial", size=12, bold=True, color=COLOR_INK)
FONT_RETENU   = Font(name="Arial", size=13, bold=True, color=COLOR_ACCENT_RED)
FONT_HEADER_TBL = Font(name="Arial", size=9, bold=True, color=COLOR_WHITE)
FONT_NUM      = Font(name="Consolas", size=10, color=COLOR_INK)
FONT_NUM_BOLD = Font(name="Consolas", size=11, bold=True, color=COLOR_INK)
FONT_FLAG     = Font(name="Arial", size=10, color=COLOR_INK)

FILL_CONCENTRATE = PatternFill(start_color=COLOR_CONCENTRATE, end_color=COLOR_CONCENTRATE, fill_type="solid")
FILL_ISOLATE     = PatternFill(start_color=COLOR_ISOLATE, end_color=COLOR_ISOLATE, fill_type="solid")
FILL_RECOVERY    = PatternFill(start_color=COLOR_RECOVERY, end_color=COLOR_RECOVERY, fill_type="solid")
FILL_MUTED       = PatternFill(start_color=COLOR_MUTED, end_color=COLOR_MUTED, fill_type="solid")
FILL_LIGHT       = PatternFill(start_color=COLOR_LIGHT, end_color=COLOR_LIGHT, fill_type="solid")
FILL_AMBER       = PatternFill(start_color=COLOR_AMBER, end_color=COLOR_AMBER, fill_type="solid")
FILL_AMBER_STRONG = PatternFill(start_color=COLOR_AMBER_STRONG, end_color=COLOR_AMBER_STRONG, fill_type="solid")
FILL_WHITE       = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

BORDER_BOTTOM = Border(bottom=Side(style="thin", color=COLOR_BORDER))

NUM_EUR = '#,##0.00" €"'
NUM_PCT = "0.0%"
NUM_PCT_SIGNED = '+0.0%;-0.0%;0.0%'

# Benchmark references (use existing Synthèse prix cells — same as v5)
CONCENTRATE_BENCH = [
    "'Synthèse prix'!H14", "'Synthèse prix'!H15", "'Synthèse prix'!H19",
    "'Synthèse prix'!H20", "'Synthèse prix'!H24", "'Synthèse prix'!H25",
    "'Synthèse prix'!H26", "'Synthèse prix'!H31", "'Synthèse prix'!H37",
    "'Synthèse prix'!H50",
]

ISOLATE_BENCH = [
    "'Synthèse prix'!H7", "'Synthèse prix'!H10", "'Synthèse prix'!H11",
    "'Synthèse prix'!H18", "'Synthèse prix'!H21", "'Synthèse prix'!H22",
    "'Synthèse prix'!H23", "'Synthèse prix'!H29", "'Synthèse prix'!H30",
    "'Synthèse prix'!H35", "'Synthèse prix'!H41", "'Synthèse prix'!H47",
]

RECOVERY_BENCH = [
    "'Synthèse prix'!H5",  # Impulse Recovery actuel (chocolat)
    "'Synthèse prix'!H6",  # Impulse Recovery actuel (vanille)
    "'Synthèse prix'!H32", # EAFIT Mega Whey (if present)
]

# Option offsets for dynamic tables
OFFSETS = [-4, -3, -2, -1, 0, 1, 2, 3, 5, 7]

# ─── Helpers ───────────────────────────────────────────────────────────────
def apply_section_header(ws, row, label, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=label)
    cell.font = FONT_SECTION
    cell.fill = fill
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 30


def write_dropdown(ws, cell_ref, default=5):
    ws[cell_ref] = default
    ws[cell_ref].fill = FILL_AMBER_STRONG
    ws[cell_ref].font = Font(name="Arial", size=14, bold=True, color=COLOR_INK)
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
    dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_ref)


def write_retenu_row(ws, row, sel_cell, options_col_letter, options_start, options_end,
                     margin_col_letter, concentrate_retenu_cell=None):
    """Write the row that shows : Option sélectionnée | Prix retenu | Marge."""
    ws.cell(row=row, column=1, value="🎯 Option sélectionnée").font = FONT_LABEL
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
    # B{row} = dropdown
    write_dropdown(ws, f"B{row}")
    ws.row_dimensions[row].height = 26

    # D{row} = Prix retenu
    ws.cell(row=row, column=3, value="→  Prix retenu").font = FONT_LABEL
    ws.cell(row=row, column=3).alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=row, column=4,
            value=f"=INDEX(${options_col_letter}${options_start}:${options_col_letter}${options_end},$B${row})")
    ws.cell(row=row, column=4).number_format = NUM_EUR
    ws.cell(row=row, column=4).font = FONT_RETENU
    ws.cell(row=row, column=4).fill = FILL_AMBER
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="center", vertical="center")

    # F{row} = Marge
    ws.cell(row=row, column=5, value="Marge nette").font = FONT_LABEL
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=row, column=6,
            value=f"=INDEX(${margin_col_letter}${options_start}:${margin_col_letter}${options_end},$B${row})")
    ws.cell(row=row, column=6).number_format = NUM_PCT
    ws.cell(row=row, column=6).font = FONT_RETENU
    ws.cell(row=row, column=6).fill = FILL_AMBER
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="center", vertical="center")

    # Optional: écart vs concentrate
    if concentrate_retenu_cell:
        ws.cell(row=row, column=7, value="vs Concentrate").font = FONT_LABEL
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="right", indent=1)
        ws.cell(row=row, column=8,
                value=f"=IFERROR((D{row}-{concentrate_retenu_cell})/{concentrate_retenu_cell},0)")
        ws.cell(row=row, column=8).number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=8).font = FONT_VALUE
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")


def write_bench_stats(ws, row, bench_refs, label="Benchmark concurrence"):
    """Write benchmark stats (avg, median, min, max) on a single row."""
    bench_list = ",".join(bench_refs)
    ws.cell(row=row, column=1, value=label).font = FONT_LABEL_MUTED
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)

    ws.cell(row=row, column=2, value="Moyenne").font = FONT_LABEL_MUTED
    ws.cell(row=row, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=3, value=f"=AVERAGE({bench_list})")
    ws.cell(row=row, column=3).number_format = NUM_EUR
    ws.cell(row=row, column=3).font = FONT_NUM

    ws.cell(row=row, column=4, value="Médiane").font = FONT_LABEL_MUTED
    ws.cell(row=row, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5, value=f"=MEDIAN({bench_list})")
    ws.cell(row=row, column=5).number_format = NUM_EUR
    ws.cell(row=row, column=5).font = FONT_NUM

    ws.cell(row=row, column=6, value="Min").font = FONT_LABEL_MUTED
    ws.cell(row=row, column=6).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=7, value=f"=MIN({bench_list})")
    ws.cell(row=row, column=7).number_format = NUM_EUR
    ws.cell(row=row, column=7).font = FONT_NUM

    ws.cell(row=row, column=8, value="Max").font = FONT_LABEL_MUTED
    ws.cell(row=row, column=8).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=9, value=f"=MAX({bench_list})")
    ws.cell(row=row, column=9).number_format = NUM_EUR
    ws.cell(row=row, column=9).font = FONT_NUM

    return f"E{row}"  # return the median cell ref


def write_table_headers(ws, row, has_concentrate_col=False):
    headers = [
        "#", "PP TTC", "PP HT", "PVM net", "Marge €", "% Marge",
        "vs Médiane",
        "vs Concentrate" if has_concentrate_col else "—",
        "Recommandation",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = FONT_HEADER_TBL
        c.fill = FILL_MUTED
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 22


def write_option_row(ws, row, option_num, pp_ttc_expr, pr_cell, median_cell,
                     concentrate_cell=None, anti_cannib_threshold=None,
                     is_concentrate_table=False):
    """Write a single option row with all the derived columns."""
    # Column A : #
    c = ws.cell(row=row, column=1, value=option_num)
    c.font = Font(name="Arial", size=10, bold=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="center")

    # Column B : PP TTC
    ws.cell(row=row, column=2, value=pp_ttc_expr)
    ws.cell(row=row, column=2).number_format = NUM_EUR

    # Column C : PP HT
    ws.cell(row=row, column=3, value=f"=B{row}/(1+$B$5)")
    ws.cell(row=row, column=3).number_format = NUM_EUR

    # Column D : PVM net
    ws.cell(row=row, column=4, value=f"=B{row}*(1-$B$6)")
    ws.cell(row=row, column=4).number_format = NUM_EUR

    # Column E : Marge €
    ws.cell(row=row, column=5, value=f"=D{row}-{pr_cell}")
    ws.cell(row=row, column=5).number_format = NUM_EUR

    # Column F : % Marge
    ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/D{row},0)")
    ws.cell(row=row, column=6).number_format = NUM_PCT

    # Column G : vs Médiane
    ws.cell(row=row, column=7, value=f"=IFERROR((B{row}-{median_cell})/{median_cell},0)")
    ws.cell(row=row, column=7).number_format = NUM_PCT_SIGNED

    # Column H : vs Concentrate (empty for Concentrate table)
    if concentrate_cell:
        ws.cell(row=row, column=8, value=f"=IFERROR((B{row}-{concentrate_cell})/{concentrate_cell},0)")
        ws.cell(row=row, column=8).number_format = NUM_PCT_SIGNED
    else:
        ws.cell(row=row, column=8, value="—")
        ws.cell(row=row, column=8).alignment = Alignment(horizontal="center")

    # Column I : Recommandation
    if is_concentrate_table:
        flag_formula = (
            f'=IF(F{row}<0.30,"⚠️ Marge basse",'
            f'IF(F{row}<0.35,"🟡 Limite",'
            f'IF(F{row}<=0.42,"🎯 Optimal","🔵 Premium")))'
        )
    else:
        cannib = anti_cannib_threshold or 0.10
        flag_formula = (
            f'=IF(F{row}<0.30,"⚠️ Marge basse",'
            f'IF(H{row}<{cannib},"🚫 Cannibalisme",'
            f'IF(F{row}<0.35,"🟡 Limite",'
            f'IF(F{row}<=0.42,"🎯 Optimal","🔵 Premium"))))'
        )
    ws.cell(row=row, column=9, value=flag_formula)
    ws.cell(row=row, column=9).font = FONT_FLAG

    # Apply fonts + alignment to numeric cols
    for col in range(2, 9):
        cell = ws.cell(row=row, column=col)
        if not cell.font or cell.font.name != "Consolas":
            cell.font = FONT_NUM
        cell.alignment = Alignment(horizontal="right", indent=1)

    # Zebra striping (alternating rows)
    if option_num % 2 == 0:
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color or cell.fill.start_color.value in (None, "00000000"):
                cell.fill = FILL_LIGHT


# ─── Table builders ────────────────────────────────────────────────────────
def build_concentrate_table(ws, start_row, hyp_cells):
    """Build the Concentrate table. Returns dict with next_row + retenu_cell."""
    r = start_row
    apply_section_header(ws, r, "▌ TABLEAU 1 — WHEY CONCENTRATE", FILL_CONCENTRATE)
    r += 2  # skip 1

    # Benchmark stats row
    bench_row = r
    median_cell = write_bench_stats(ws, r, CONCENTRATE_BENCH)
    r += 2

    # Headers
    headers_row = r
    write_table_headers(ws, r, has_concentrate_col=False)
    r += 1

    # 10 options (hardcoded prices)
    concentrate_prices = [29.9, 31.9, 33.9, 34.9, 36.9, 37.9, 39.9, 41.9, 42.9, 44.9]
    options_start = r
    pr_cell = hyp_cells["PR_conc"]
    for i, price in enumerate(concentrate_prices, 1):
        write_option_row(
            ws, r, i, price, pr_cell, median_cell,
            concentrate_cell=None, is_concentrate_table=True,
        )
        r += 1
    options_end = r - 1

    # Sélection row (placed just below the table)
    r += 1
    sel_row = r
    write_retenu_row(ws, r, f"B{r}", "B", options_start, options_end, "F")
    retenu_cell = f"D{r}"
    margin_retenu_cell = f"F{r}"

    return {
        "next_row": r + 2,
        "retenu_cell": retenu_cell,
        "margin_cell": margin_retenu_cell,
        "median_cell": median_cell,
    }


def build_isolate_table(ws, start_row, hyp_cells, concentrate_retenu_cell):
    """Build the Isolate table (dynamic). Returns dict with next_row + retenu_cell."""
    r = start_row
    apply_section_header(ws, r, "▌ TABLEAU 2 — WHEY ISOLATE (auto-calé sur Concentrate)", FILL_ISOLATE)
    r += 2

    # Base + calcul auto block
    base_row = r
    ws.cell(row=r, column=1, value="Base cascade").font = FONT_LABEL_MUTED
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=2, value="Concentrate retenu").font = FONT_LABEL_MUTED
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=3, value=f"={concentrate_retenu_cell}").number_format = NUM_EUR
    ws.cell(row=r, column=3).font = FONT_NUM_BOLD
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    r += 2

    # Calcul auto : 4 bornes + PP optimal
    calc_row = r
    # Labels
    labels = [
        (1, "Plancher 38% marge"),
        (3, "Cible 40% marge"),
        (5, "Anti-cannib (C+10%)"),
        (7, "Marché médiane"),
    ]
    for col, lbl in labels:
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = FONT_LABEL_MUTED
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=9, value="→ PP OPTIMAL").font = Font(name="Arial", size=10, bold=True, color=COLOR_ACCENT_RED)
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")
    r += 1

    # Values
    values_row = r
    pr_iso = hyp_cells["PR_iso"]
    remise = hyp_cells["remise"]
    bench_list = ",".join(ISOLATE_BENCH)
    # Plancher 38% : PR / 0.62 / (1-remise), rounded up to X.9
    ws.cell(row=r, column=2, value=f"=CEILING({pr_iso}/(1-0.38)/(1-{remise})+0.1,1)-0.1")
    marge38_cell = f"B{r}"
    # Cible 40%
    ws.cell(row=r, column=4, value=f"=CEILING({pr_iso}/(1-0.40)/(1-{remise})+0.1,1)-0.1")
    marge40_cell = f"D{r}"
    # Anti-cannib = Concentrate × 1.10 rounded up
    ws.cell(row=r, column=6, value=f"=CEILING({concentrate_retenu_cell}*1.10+0.1,1)-0.1")
    anticannib_cell = f"F{r}"
    # Marché médiane (live)
    ws.cell(row=r, column=8, value=f"=MEDIAN({bench_list})")
    marche_cell = f"H{r}"
    # PP OPTIMAL
    pp_opt_formula = f"=MAX({marge38_cell},{anticannib_cell},MIN({marge40_cell},{marche_cell}))"
    ws.cell(row=r, column=9, value=pp_opt_formula)
    pp_opt_cell = f"I{r}"
    # Formatting
    for col in [2, 4, 6, 8]:
        ws.cell(row=r, column=col).number_format = NUM_EUR
        ws.cell(row=r, column=col).font = FONT_NUM
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=9).number_format = NUM_EUR
    ws.cell(row=r, column=9).font = Font(name="Arial", size=13, bold=True, color=COLOR_WHITE)
    ws.cell(row=r, column=9).fill = PatternFill(start_color=COLOR_ACCENT_RED, end_color=COLOR_ACCENT_RED, fill_type="solid")
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
    r += 2

    # Benchmark stats row
    median_cell = write_bench_stats(ws, r, ISOLATE_BENCH)
    r += 2

    # Headers
    write_table_headers(ws, r, has_concentrate_col=True)
    r += 1

    # 10 options (dynamic formulas)
    options_start = r
    for i, offset in enumerate(OFFSETS, 1):
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"{pp_opt_cell}" if offset == 0 else f"{pp_opt_cell}{sign}{abs_off}"
        pp_ttc_formula = f"=CEILING({expr}+0.1,1)-0.1"
        write_option_row(
            ws, r, i, pp_ttc_formula, pr_iso, median_cell,
            concentrate_cell=concentrate_retenu_cell,
            anti_cannib_threshold=0.10,
            is_concentrate_table=False,
        )
        r += 1
    options_end = r - 1

    # Sélection row
    r += 1
    write_retenu_row(ws, r, f"B{r}", "B", options_start, options_end, "F",
                     concentrate_retenu_cell=concentrate_retenu_cell)
    retenu_cell = f"D{r}"

    return {"next_row": r + 2, "retenu_cell": retenu_cell}


def build_recovery_table(ws, start_row, hyp_cells, concentrate_retenu_cell, isolate_retenu_cell):
    r = start_row
    apply_section_header(ws, r, "▌ TABLEAU 3 — WHEY RECOVERY (auto-calé sur gamme)", FILL_RECOVERY)
    r += 2

    # Base : dual cascade
    ws.cell(row=r, column=1, value="Base cascade").font = FONT_LABEL_MUTED
    ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", indent=1)
    ws.cell(row=r, column=2, value="Concentrate retenu").font = FONT_LABEL_MUTED
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=3, value=f"={concentrate_retenu_cell}").number_format = NUM_EUR
    ws.cell(row=r, column=3).font = FONT_NUM_BOLD
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=4, value="Isolate retenu").font = FONT_LABEL_MUTED
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="right")
    ws.cell(row=r, column=5, value=f"={isolate_retenu_cell}").number_format = NUM_EUR
    ws.cell(row=r, column=5).font = FONT_NUM_BOLD
    ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
    r += 2

    # Calcul auto
    labels = [
        (1, "Plancher 38% marge"),
        (3, "Plancher C+5%"),
        (5, "Plafond I−5%"),
        (7, "Marché réf (Impulse)"),
    ]
    for col, lbl in labels:
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = FONT_LABEL_MUTED
        c.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=9, value="→ PP OPTIMAL").font = Font(name="Arial", size=10, bold=True, color=COLOR_ACCENT_RED)
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")
    r += 1

    pr_rec = hyp_cells["PR_rec"]
    remise = hyp_cells["remise"]
    ws.cell(row=r, column=2, value=f"=CEILING({pr_rec}/(1-0.38)/(1-{remise})+0.1,1)-0.1")
    marge38_cell = f"B{r}"
    ws.cell(row=r, column=4, value=f"=CEILING({concentrate_retenu_cell}*1.05+0.1,1)-0.1")
    floor_cell = f"D{r}"
    ws.cell(row=r, column=6, value=f"=CEILING({isolate_retenu_cell}*0.95+0.1,1)-0.1")
    ceiling_cell = f"F{r}"
    # Marché réf : Impulse current whey recovery (from Synthèse prix H5)
    ws.cell(row=r, column=8, value="='Synthèse prix'!H5")
    marche_cell = f"H{r}"
    pp_opt_formula = f"=MAX({marge38_cell},{floor_cell},MIN({ceiling_cell},{marche_cell}))"
    ws.cell(row=r, column=9, value=pp_opt_formula)
    pp_opt_cell = f"I{r}"
    for col in [2, 4, 6, 8]:
        ws.cell(row=r, column=col).number_format = NUM_EUR
        ws.cell(row=r, column=col).font = FONT_NUM
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=9).number_format = NUM_EUR
    ws.cell(row=r, column=9).font = Font(name="Arial", size=13, bold=True, color=COLOR_WHITE)
    ws.cell(row=r, column=9).fill = PatternFill(start_color=COLOR_ACCENT_RED, end_color=COLOR_ACCENT_RED, fill_type="solid")
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
    r += 2

    # Benchmark stats
    median_cell = write_bench_stats(ws, r, RECOVERY_BENCH, label="Benchmark Recovery (réduit)")
    r += 2

    # Headers
    write_table_headers(ws, r, has_concentrate_col=True)
    r += 1

    options_start = r
    for i, offset in enumerate(OFFSETS, 1):
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"{pp_opt_cell}" if offset == 0 else f"{pp_opt_cell}{sign}{abs_off}"
        pp_ttc_formula = f"=CEILING({expr}+0.1,1)-0.1"
        write_option_row(
            ws, r, i, pp_ttc_formula, pr_rec, median_cell,
            concentrate_cell=concentrate_retenu_cell,
            anti_cannib_threshold=0.05,
            is_concentrate_table=False,
        )
        r += 1
    options_end = r - 1

    r += 1
    write_retenu_row(ws, r, f"B{r}", "B", options_start, options_end, "F",
                     concentrate_retenu_cell=concentrate_retenu_cell)
    retenu_cell = f"D{r}"

    return {"next_row": r + 2, "retenu_cell": retenu_cell}


# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    # Remove the old pricing tabs from v5 since the user wants a single tab
    for old_name in ("Pricing Whey Concentrate", "Pricing Whey Isolate", "Pricing Whey Recovery"):
        if old_name in wb.sheetnames:
            del wb[old_name]

    # Create new tab at position 0
    if "Pricing Dynamique" in wb.sheetnames:
        del wb["Pricing Dynamique"]
    ws = wb.create_sheet("Pricing Dynamique", 0)

    # ─── Title ───
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "PRICING DYNAMIQUE — GAMME WHEY IMPULSE"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "1 sélection Concentrate → Isolate & Recovery se recalculent automatiquement"
    c.font = FONT_SUBTITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ─── Hypothèses (rows 4-9) ───
    c = ws.cell(row=4, column=1, value="HYPOTHÈSES PARTAGÉES")
    c.font = Font(name="Arial", size=10, bold=True, color=COLOR_MUTED)
    c.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("A4:I4")
    ws.cell(row=4, column=1).border = BORDER_BOTTOM

    rows_hyp = [
        (5, "TVA (compléments alim.)", 0.055, NUM_PCT),
        (6, "Remise client moyenne", 0.18, NUM_PCT),
        (7, "PR Concentrate (€ HT)", 21, NUM_EUR),
        (8, "PR Isolate (€ HT)", 25, NUM_EUR),
        (9, "PR Recovery (€ HT)", 12.15, NUM_EUR),
    ]
    for row, label, value, num_fmt in rows_hyp:
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = num_fmt
        c.font = FONT_NUM_BOLD
        c.alignment = Alignment(horizontal="center")
        c.fill = FILL_LIGHT

    hyp_cells = {
        "TVA": "$B$5",
        "remise": "$B$6",
        "PR_conc": "$B$7",
        "PR_iso": "$B$8",
        "PR_rec": "$B$9",
    }

    # ─── Build the 3 tables ───
    concentrate = build_concentrate_table(ws, start_row=11, hyp_cells=hyp_cells)
    isolate = build_isolate_table(
        ws, start_row=concentrate["next_row"] + 1,
        hyp_cells=hyp_cells,
        concentrate_retenu_cell=concentrate["retenu_cell"],
    )
    recovery = build_recovery_table(
        ws, start_row=isolate["next_row"] + 1,
        hyp_cells=hyp_cells,
        concentrate_retenu_cell=concentrate["retenu_cell"],
        isolate_retenu_cell=isolate["retenu_cell"],
    )

    # ─── Column widths ───
    widths = {"A": 26, "B": 13, "C": 13, "D": 14, "E": 12, "F": 13, "G": 16, "H": 18, "I": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze panes below hypothèses
    ws.freeze_panes = "A11"

    # Save
    wb.save(DST)
    print(f"✓ Generated {DST}")
    print(f"  Concentrate retenu : {concentrate['retenu_cell']}")
    print(f"  Isolate retenu     : {isolate['retenu_cell']}")
    print(f"  Recovery retenu    : {recovery['retenu_cell']}")


if __name__ == "__main__":
    main()
