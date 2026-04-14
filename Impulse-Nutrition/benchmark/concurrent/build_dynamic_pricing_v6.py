"""
Build benchmark_whey_concurrent_v6_dynamic.xlsx

Takes the user's existing xlsx (v5 built with his colleague) and refactors
the "Pricing Whey Isolate" and "Pricing Whey Recovery" tabs so that the 10
price options are computed dynamically from:
  - PR_product (fixed cell)
  - TVA (linked from Data)
  - Remise client (linked from Data)
  - Concentrate_retenu (linked from Pricing Whey Concentrate tab)
  - Market median / benchmark (already computed in current tab)

Zero hardcoded parameters. Everything cascades from Concentrate_retenu.

All other tabs (Benchmark détaillé, Synthèse prix, Discounts, Classement,
Data, Pricing Whey Concentrate) are left untouched.
"""
import shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = Path("/Users/antoinechabrat/Downloads/Copie de benchmark_whey_concurrent 5 1.xlsx")
DST = Path(
    "/Users/antoinechabrat/Documents/SmallProject/Impulse-Nutrition/benchmark/"
    "concurrent/benchmark_whey_concurrent_v6_dynamic.xlsx"
)

SECTION_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SECTION_FONT = Font(color="FFFFFF", bold=True, size=11)
LABEL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
LABEL_FONT = Font(bold=True, size=10)
FORMULA_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
FORMULA_FONT = Font(italic=True, color="1F3864", size=10)
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

# Round-up-to-X.9 formula snippet
R09 = "CEILING({expr}+0.1,1)-0.1"


def round09(expr: str) -> str:
    return R09.format(expr=expr)


def build_isolate(ws):
    """Refactor Pricing Whey Isolate tab with dynamic Option C."""

    # ── 1. Add CALCUL AUTO block at rows 38-47
    rows = [
        (38, "A", "CALCUL AUTO — PRIX OPTIMAL (ALGO OPTION C)", SECTION_FONT, SECTION_FILL),
        (39, "A", "PP pour marge 38% (plancher rentabilité)", LABEL_FONT, LABEL_FILL),
        (39, "B", f"={round09('$B$5/(1-0.38)/(1-$B$7)')}", FORMULA_FONT, FORMULA_FILL),
        (40, "A", "PP pour marge 40% (cible marge)", LABEL_FONT, LABEL_FILL),
        (40, "B", f"={round09('$B$5/(1-0.40)/(1-$B$7)')}", FORMULA_FONT, FORMULA_FILL),
        (41, "A", "PP plancher anti-cannib. (Concentrate +10%)", LABEL_FONT, LABEL_FILL),
        (41, "B", f"={round09('$B$11*1.10')}", FORMULA_FONT, FORMULA_FILL),
        (42, "A", "PP médiane marché (benchmark concurrents)", LABEL_FONT, LABEL_FILL),
        (42, "B", "=$B$28", FORMULA_FONT, FORMULA_FILL),
        (43, "A", "🎯 PP OPTIMAL (auto)", SECTION_FONT, SECTION_FILL),
        (
            43, "B",
            "=MAX($B$39,$B$41,MIN($B$40,$B$42))",
            Font(bold=True, color="FFFFFF", size=12),
            SECTION_FILL,
        ),
        (44, "A", "Marge théorique à PP optimal", LABEL_FONT, LABEL_FILL),
        (44, "B", "=IFERROR(($B$43*(1-$B$7)-$B$5)/($B$43*(1-$B$7)),0)", FORMULA_FONT, FORMULA_FILL),
        (45, "A", "Écart vs Concentrate retenu (€)", LABEL_FONT, LABEL_FILL),
        (45, "B", "=$B$43-$B$11", FORMULA_FONT, FORMULA_FILL),
        (46, "A", "Écart vs Concentrate retenu (%)", LABEL_FONT, LABEL_FILL),
        (46, "B", "=IFERROR(($B$43-$B$11)/$B$11,0)", FORMULA_FONT, FORMULA_FILL),
        (47, "A", "Écart vs médiane marché (€)", LABEL_FONT, LABEL_FILL),
        (47, "B", "=$B$43-$B$42", FORMULA_FONT, FORMULA_FILL),
    ]
    for row, col, value, font, fill in rows:
        c = ws[f"{col}{row}"]
        c.value = value
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="left" if col == "A" else "right", vertical="center")
    # Percentage format for margin cell
    ws["B44"].number_format = "0.0%"
    ws["B46"].number_format = "0.0%"
    # Column widths
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 30, 44)

    # ── 2. Replace E4:E13 with dynamic formulas referencing B43 (PP_optimal)
    offsets = [-4, -3, -2, -1, 0, 1, 2, 3, 5, 7]
    for i, offset in enumerate(offsets):
        row = 4 + i
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"$B$43{sign}{abs_off}" if offset != 0 else "$B$43"
        formula = f"={round09(expr)}"
        cell = ws[f"E{row}"]
        cell.value = formula
        cell.font = FORMULA_FONT
        cell.fill = FORMULA_FILL
    # Highlight center row (E8 = PP optimal)
    ws["E8"].fill = HIGHLIGHT_FILL
    ws["E8"].font = Font(bold=True, color="C00000", size=11)


def build_recovery(ws):
    """Refactor Pricing Whey Recovery tab with dynamic Option C."""

    # ── 1. Add CALCUL AUTO block at rows 20-30
    rows = [
        (20, "A", "CALCUL AUTO — PRIX OPTIMAL (ALGO OPTION C)", SECTION_FONT, SECTION_FILL),
        (21, "A", "PP pour marge 38% (plancher rentabilité)", LABEL_FONT, LABEL_FILL),
        (21, "B", f"={round09('$B$5/(1-0.38)/(1-$B$7)')}", FORMULA_FONT, FORMULA_FILL),
        (22, "A", "PP pour marge 40% (cible marge)", LABEL_FONT, LABEL_FILL),
        (22, "B", f"={round09('$B$5/(1-0.40)/(1-$B$7)')}", FORMULA_FONT, FORMULA_FILL),
        (23, "A", "PP plancher anti-cannib. (Concentrate +5%)", LABEL_FONT, LABEL_FILL),
        (23, "B", f"={round09('$B$11*1.05')}", FORMULA_FONT, FORMULA_FILL),
        (24, "A", "PP plafond anti-cannib. (Isolate −5%)", LABEL_FONT, LABEL_FILL),
        (24, "B", f"={round09('$B$12*0.95')}", FORMULA_FONT, FORMULA_FILL),
        (25, "A", "PP référence marché (Impulse actuel)", LABEL_FONT, LABEL_FILL),
        (25, "B", "=$B$8", FORMULA_FONT, FORMULA_FILL),
        (26, "A", "🎯 PP OPTIMAL (auto)", SECTION_FONT, SECTION_FILL),
        (
            26, "B",
            "=MAX($B$21,$B$23,MIN($B$25,$B$24))",
            Font(bold=True, color="FFFFFF", size=12),
            SECTION_FILL,
        ),
        (27, "A", "Marge théorique à PP optimal", LABEL_FONT, LABEL_FILL),
        (27, "B", "=IFERROR(($B$26*(1-$B$7)-$B$5)/($B$26*(1-$B$7)),0)", FORMULA_FONT, FORMULA_FILL),
        (28, "A", "Écart vs Concentrate retenu (€)", LABEL_FONT, LABEL_FILL),
        (28, "B", "=$B$26-$B$11", FORMULA_FONT, FORMULA_FILL),
        (29, "A", "Écart vs Isolate retenu (€)", LABEL_FONT, LABEL_FILL),
        (29, "B", "=$B$26-$B$12", FORMULA_FONT, FORMULA_FILL),
    ]
    for row, col, value, font, fill in rows:
        c = ws[f"{col}{row}"]
        c.value = value
        c.font = font
        c.fill = fill
        c.alignment = Alignment(horizontal="left" if col == "A" else "right", vertical="center")
    ws["B27"].number_format = "0.0%"
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 30, 44)

    # ── 2. Add PRIX RETENU block at rows 31-35 (doesn't exist yet in Recovery)
    retenu_rows = [
        (31, "A", "PRIX RETENU RECOVERY", SECTION_FONT, SECTION_FILL),
        (32, "A", "PP TTC retenu (🎯 Optimal)", LABEL_FONT, LABEL_FILL),
        (
            32, "B",
            "=IFERROR(INDEX(E4:E13,MATCH(\"🎯 Optimal\",R4:R13,0)),$B$26)",
            Font(bold=True, color="C00000", size=12),
            HIGHLIGHT_FILL,
        ),
        (33, "A", "PVM net retenu", LABEL_FONT, LABEL_FILL),
        (33, "B", "=$B$32*(1-$B$7)", FORMULA_FONT, FORMULA_FILL),
        (34, "A", "% Marge nette retenu", LABEL_FONT, LABEL_FILL),
        (34, "B", "=IFERROR(($B$33-$B$5)/$B$33,0)", FORMULA_FONT, FORMULA_FILL),
        (35, "A", "Écart vs Concentrate retenu (%)", LABEL_FONT, LABEL_FILL),
        (35, "B", "=IFERROR(($B$32-$B$11)/$B$11,0)", FORMULA_FONT, FORMULA_FILL),
        (36, "A", "Écart vs Isolate retenu (%)", LABEL_FONT, LABEL_FILL),
        (36, "B", "=IFERROR(($B$32-$B$12)/$B$12,0)", FORMULA_FONT, FORMULA_FILL),
    ]
    for row, col, value, font, fill in retenu_rows:
        c = ws[f"{col}{row}"]
        c.value = value
        c.font = font
        c.fill = fill
    ws["B34"].number_format = "0.0%"
    ws["B35"].number_format = "0.0%"
    ws["B36"].number_format = "0.0%"

    # ── 3. Replace E4:E13 with dynamic formulas referencing B26 (PP_optimal)
    offsets = [-4, -3, -2, -1, 0, 1, 2, 3, 5, 7]
    for i, offset in enumerate(offsets):
        row = 4 + i
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"$B$26{sign}{abs_off}" if offset != 0 else "$B$26"
        formula = f"={round09(expr)}"
        cell = ws[f"E{row}"]
        cell.value = formula
        cell.font = FORMULA_FONT
        cell.fill = FORMULA_FILL
    ws["E8"].fill = HIGHLIGHT_FILL
    ws["E8"].font = Font(bold=True, color="C00000", size=11)


def main():
    # Copy source to dest first (preserves all existing formatting / formulas)
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)

    wb = openpyxl.load_workbook(DST)

    build_isolate(wb["Pricing Whey Isolate"])
    build_recovery(wb["Pricing Whey Recovery"])

    wb.save(DST)
    print(f"✓ Generated {DST}")


if __name__ == "__main__":
    main()
