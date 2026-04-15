"""
Build v10 — Finance research report layout.

Key differences vs v9 :
  1. NEW LAYOUT : products as columns, metrics as rows (matrice style equity
     research). P&L richer (8+ lines), benchmark inline, algo inline, all in
     the same vertical matrix. Options grid goes below, compact.
  2. NEW ALGO : hybride marché/marge — prend la médiane marché comme ancre,
     monte au plancher 30% si marché trop bas, accepte >45% si marché haut.
  3. NO BIG FILLS : subtle colors (small bullets ●), thin borders, Consolas
     for numbers. Max font 14pt (title only).
  4. Flags 3 niveaux : ⚠️ < 30% / 🎯 30-45% / 🔵 > 45%, plus 🚫 cannibalisme.
  5. P&L detailed : 8 lines (TTC, -TVA, PP HT, -remise, PVM net, -PR, marge €, % marge).
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path("/Users/antoinechabrat/Downloads/Copie de benchmark_whey_concurrent 5 1.xlsx")
DST = Path(
    "/Users/antoinechabrat/Documents/SmallProject/Impulse-Nutrition/benchmark/"
    "concurrent/benchmark_whey_concurrent_v10_dynamic.xlsx"
)

# ─── Design tokens — SOBRE FINANCE ─────────────────────────────────────
INK       = "0F172A"  # main text
MUTED     = "64748B"  # secondary text
SUBTLE    = "94A3B8"  # tertiary text
BORDER    = "E2E8F0"  # thin border
BORDER_STRONG = "CBD5E1"
LIGHT     = "F8FAFC"  # zebra
WHITE     = "FFFFFF"
AMBER_SOFT = "FEF3C7"  # editable cells
AMBER_STR  = "F59E0B"  # dropdown

# Product accent dots (text color, not fill)
DOT_CONC = "1D4ED8"  # blue-700
DOT_ISO  = "047857"  # emerald-700
DOT_REC  = "B91C1C"  # red-700

# Flag colors (text only)
FLAG_WARN = "B91C1C"
FLAG_OPT  = "047857"
FLAG_PREM = "1D4ED8"
FLAG_CANNIB = "DC2626"

FONT_TITLE    = Font(name="Arial", size=14, bold=True, color=INK)
FONT_SUB      = Font(name="Arial", size=9, italic=True, color=MUTED)
FONT_SECTION  = Font(name="Arial", size=8, bold=True, color=MUTED)  # UPPERCASE labels
FONT_LABEL    = Font(name="Arial", size=9, color=INK)
FONT_LABEL_MUT = Font(name="Arial", size=9, italic=True, color=MUTED)
FONT_LABEL_B  = Font(name="Arial", size=9, bold=True, color=INK)
FONT_NUM      = Font(name="Consolas", size=9, color=INK)
FONT_NUM_MUTED= Font(name="Consolas", size=9, color=MUTED)
FONT_NUM_B    = Font(name="Consolas", size=10, bold=True, color=INK)
FONT_NUM_ITAL = Font(name="Consolas", size=9, italic=True, color=INK)
FONT_NUM_NEG  = Font(name="Consolas", size=9, color=FLAG_WARN)
FONT_NUM_RET  = Font(name="Arial", size=11, bold=True, color=INK)
FONT_HEADER_PROD = {
    "conc": Font(name="Arial", size=9, bold=True, color=DOT_CONC),
    "iso":  Font(name="Arial", size=9, bold=True, color=DOT_ISO),
    "rec":  Font(name="Arial", size=9, bold=True, color=DOT_REC),
}
FONT_DROPDOWN = Font(name="Arial", size=10, bold=True, color=INK)
FONT_FLAG_SIZE = Font(name="Arial", size=9, color=INK)

FILL_WHITE   = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT   = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
FILL_AMBER   = PatternFill(start_color=AMBER_SOFT, end_color=AMBER_SOFT, fill_type="solid")
FILL_AMBER_S = PatternFill(start_color=AMBER_STR, end_color=AMBER_STR, fill_type="solid")

THIN_GREY   = Side(style="thin", color=BORDER_STRONG)
THIN_INK    = Side(style="thin", color=INK)
DOTTED      = Side(style="dotted", color=BORDER_STRONG)

BOX   = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
TOP   = Border(top=THIN_GREY)
TOP_INK = Border(top=THIN_INK)
BOT_INK = Border(bottom=THIN_INK)
TOP_BOT = Border(top=THIN_INK, bottom=THIN_INK)

NUM_EUR = '#,##0.00" €"'
NUM_EUR_NEG = '#,##0.00" €";[Red]-#,##0.00" €"'
NUM_PCT = "0.0%"
NUM_PCT_SIGNED = '+0.0%;-0.0%;0.0%'

# NET-of-discount prices: col S = "Prix client standard" = Prix actuel × (1 - remise moyenne marque)
# Source: Synthèse prix!S = IFERROR(I*(1-Q/100),"") where Q = VLOOKUP marque dans "Discounts par marque"
# Homogène avec Impulse qui a aussi une remise client moyenne de ~18%.
CONC_BENCH = [
    "'Synthèse prix'!S14", "'Synthèse prix'!S15", "'Synthèse prix'!S19",
    "'Synthèse prix'!S20", "'Synthèse prix'!S24", "'Synthèse prix'!S25",
    "'Synthèse prix'!S26", "'Synthèse prix'!S31", "'Synthèse prix'!S37",
    "'Synthèse prix'!S50",
]
ISO_BENCH = [
    "'Synthèse prix'!S7", "'Synthèse prix'!S10", "'Synthèse prix'!S11",
    "'Synthèse prix'!S18", "'Synthèse prix'!S21", "'Synthèse prix'!S22",
    "'Synthèse prix'!S23", "'Synthèse prix'!S29", "'Synthèse prix'!S30",
    "'Synthèse prix'!S35", "'Synthèse prix'!S41", "'Synthèse prix'!S47",
]
REC_BENCH = [
    "'Synthèse prix'!S5", "'Synthèse prix'!S6", "'Synthèse prix'!S32",
]

# Option 1 = pivot éditable (prix actuel Impulse par produit)
# Options 2→10 cascadent en +1, +2, +3, +4, +5, +6, +8, +10, +12 € depuis le pivot
OFFSETS = [0, 1, 2, 3, 4, 5, 6, 8, 10, 12]
PIVOT_DEFAULT = 39.90

# Column layout (13 cols)
# A : label (width 30)
# B : Concentrate "Prix" (in options) / merged value (in P&L)
# C : Concentrate "Marge" (options only)
# D : Concentrate "Flag" (options only)
# E : gap
# F-H : Isolate
# I : gap
# J-L : Recovery
# M : optional Δ gamme
COL_LBL      = 1
COL_C_PRIX   = 2
COL_C_MARGE  = 3
COL_C_FLAG   = 4
COL_I_PRIX   = 6
COL_I_MARGE  = 7
COL_I_FLAG   = 8
COL_R_PRIX   = 10
COL_R_MARGE  = 11
COL_R_FLAG   = 12
COL_DELTA    = 13


def add_dropdown(ws, cell_ref):
    dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_ref)
    ws[cell_ref] = 1  # défaut = option 1 (= pivot, prix actuel)
    ws[cell_ref].font = FONT_DROPDOWN
    ws[cell_ref].fill = FILL_AMBER_S
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell_ref].border = BOX


def write_section_header(ws, row, label):
    """Thin-line section header: label in small uppercase grey + top border."""
    ws.cell(row=row, column=COL_LBL, value=label).font = FONT_SECTION
    ws.cell(row=row, column=COL_LBL).alignment = Alignment(horizontal="left", indent=0)
    # Apply top border across the row
    for col in range(1, COL_DELTA + 1):
        ws.cell(row=row, column=col).border = TOP
    ws.row_dimensions[row].height = 16


def write_pnl_row(ws, row, label, formula_builder, retenus, italic=False, bold=False, top_border=False, is_pct=False, is_negative=False):
    """Write one P&L-style row. formula_builder(retenu_cell, pr_cell) returns the formula string.

    retenus : dict {product: (retenu_cell, pr_cell)}
    """
    # Label
    lbl = ws.cell(row=row, column=COL_LBL, value=label)
    if bold:
        lbl.font = FONT_LABEL_B
    elif italic:
        lbl.font = Font(name="Arial", size=9, italic=True, color=INK)
    else:
        lbl.font = FONT_LABEL
    lbl.alignment = Alignment(horizontal="left", indent=2)

    products = [
        ("conc", COL_C_PRIX, COL_C_FLAG),
        ("iso",  COL_I_PRIX, COL_I_FLAG),
        ("rec",  COL_R_PRIX, COL_R_FLAG),
    ]

    for key, start_col, end_col in products:
        retenu_cell, pr_cell = retenus[key]
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        cell = ws.cell(row=row, column=start_col, value=formula_builder(retenu_cell, pr_cell))
        if is_pct:
            cell.number_format = NUM_PCT
        elif is_negative:
            cell.number_format = NUM_EUR_NEG
        else:
            cell.number_format = NUM_EUR

        if bold:
            cell.font = FONT_NUM_B
        elif italic:
            cell.font = FONT_NUM_ITAL
        elif is_negative:
            cell.font = FONT_NUM_NEG
        else:
            cell.font = FONT_NUM
        cell.alignment = Alignment(horizontal="right", indent=1)
        if top_border:
            cell.border = TOP_INK

    if top_border:
        ws.cell(row=row, column=COL_LBL).border = TOP_INK
        # Apply to all merged cells already, but also label


def build_pricing_v10(wb):
    if "Pricing Grille" in wb.sheetnames:
        del wb["Pricing Grille"]
    ws = wb.create_sheet("Pricing Grille", 0)

    # ═══ TITLE (row 1-2) ═══
    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "PRICING DYNAMIQUE  ·  GAMME WHEY IMPULSE"
    c.font = FONT_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:M2")
    c = ws["A2"]
    c.value = "v10  ·  Finance view  ·  algo hybride marché/marge  ·  HT-based"
    c.font = FONT_SUB
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ═══ HYPOTHÈSES COMMUNES (row 4-5) ═══
    write_section_header(ws, 4, "HYPOTHÈSES COMMUNES  (cellules ambre = éditables)")

    hyp_cells_row = 5
    hyp_pairs = [
        (COL_LBL, "TVA",    0.055, NUM_PCT),
        (3,       "Remise", 0.18,  NUM_PCT),
    ]
    for col, label, value, fmt in hyp_pairs:
        lbl = ws.cell(row=hyp_cells_row, column=col, value=label)
        lbl.font = FONT_LABEL_MUT
        lbl.alignment = Alignment(horizontal="right", indent=1)
        val = ws.cell(row=hyp_cells_row, column=col + 1, value=value)
        val.number_format = fmt
        val.font = FONT_NUM_B
        val.fill = FILL_AMBER
        val.alignment = Alignment(horizontal="center")
        val.border = BOX

    tva_cell    = "$B$5"
    remise_cell = "$D$5"
    # PR cells are now placed in row 10 under each product column (see below)
    pr_conc     = "$B$10"
    pr_iso      = "$F$10"
    pr_rec      = "$J$10"

    # Row 6 : marges cibles éditables (plancher / cible / plafond)
    ws.cell(row=6, column=COL_LBL, value="Marges HT cibles  (⚠️ plancher  ·  🎯 cible  ·  🔵 plafond)").font = FONT_LABEL_MUT
    ws.cell(row=6, column=COL_LBL).alignment = Alignment(horizontal="right", indent=1)

    for col, val in [(COL_C_PRIX, 0.20), (COL_C_MARGE, 0.35), (COL_C_FLAG, 0.60)]:
        c = ws.cell(row=6, column=col, value=val)
        c.number_format = NUM_PCT
        c.font = FONT_NUM_B
        c.fill = FILL_AMBER
        c.alignment = Alignment(horizontal="center")
        c.border = BOX
    ws.row_dimensions[6].height = 18

    floor_cell  = "$B$6"
    target_cell = "$C$6"
    ceil_cell   = "$D$6"

    # ═══ PRODUCT HEADER ROW (row 7) ═══
    # label empty, then product names with colored dots
    ws.cell(row=7, column=COL_LBL, value="").font = FONT_LABEL

    product_headers = [
        (COL_C_PRIX, COL_C_FLAG, "●  CONCENTRATE", DOT_CONC),
        (COL_I_PRIX, COL_I_FLAG, "●  ISOLATE", DOT_ISO),
        (COL_R_PRIX, COL_R_FLAG, "●  RECOVERY", DOT_REC),
    ]
    for start_col, end_col, label, color in product_headers:
        ws.merge_cells(start_row=7, start_column=start_col, end_row=7, end_column=end_col)
        cell = ws.cell(row=7, column=start_col, value=label)
        cell.font = Font(name="Arial", size=10, bold=True, color=color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 22
    # bottom border on header row
    for col in range(1, COL_DELTA + 1):
        ws.cell(row=7, column=col).border = BOT_INK

    # ═══ SÉLECTION + PRIX RETENU (row 8-9) ═══
    # Row 8 : "Option sélectionnée" + dropdowns
    ws.cell(row=8, column=COL_LBL, value="Option sélectionnée").font = FONT_LABEL_MUT
    ws.cell(row=8, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)

    conc_dropdown = f"{get_column_letter(COL_C_PRIX)}8"
    iso_dropdown  = f"{get_column_letter(COL_I_PRIX)}8"
    rec_dropdown  = f"{get_column_letter(COL_R_PRIX)}8"

    for cell_ref, start, end in [
        (conc_dropdown, COL_C_PRIX, COL_C_FLAG),
        (iso_dropdown, COL_I_PRIX, COL_I_FLAG),
        (rec_dropdown, COL_R_PRIX, COL_R_FLAG),
    ]:
        add_dropdown(ws, cell_ref)
        ws.merge_cells(start_row=8, start_column=start, end_row=8, end_column=end)
    ws.row_dimensions[8].height = 22

    # Row 9 : Prix retenu (bold, no fill)
    ws.cell(row=9, column=COL_LBL, value="Prix retenu (TTC)").font = FONT_LABEL_B
    ws.cell(row=9, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)

    options_start_row = 42  # where the options grid begins
    options_end_row = 51

    conc_retenu = f"{get_column_letter(COL_C_PRIX)}9"
    iso_retenu  = f"{get_column_letter(COL_I_PRIX)}9"
    rec_retenu  = f"{get_column_letter(COL_R_PRIX)}9"

    retenu_formulas = [
        (conc_retenu, COL_C_PRIX, COL_C_FLAG, f"=INDEX({get_column_letter(COL_C_PRIX)}{options_start_row}:{get_column_letter(COL_C_PRIX)}{options_end_row},{conc_dropdown})"),
        (iso_retenu,  COL_I_PRIX, COL_I_FLAG, f"=INDEX({get_column_letter(COL_I_PRIX)}{options_start_row}:{get_column_letter(COL_I_PRIX)}{options_end_row},{iso_dropdown})"),
        (rec_retenu,  COL_R_PRIX, COL_R_FLAG, f"=INDEX({get_column_letter(COL_R_PRIX)}{options_start_row}:{get_column_letter(COL_R_PRIX)}{options_end_row},{rec_dropdown})"),
    ]
    for cell_ref, start, end, formula in retenu_formulas:
        ws.merge_cells(start_row=9, start_column=start, end_row=9, end_column=end)
        c = ws.cell(row=9, column=start, value=formula)
        c.font = FONT_NUM_RET
        c.number_format = NUM_EUR
        c.alignment = Alignment(horizontal="right", indent=1)
    ws.row_dimensions[9].height = 20

    retenus = {
        "conc": (conc_retenu, pr_conc),
        "iso":  (iso_retenu, pr_iso),
        "rec":  (rec_retenu, pr_rec),
    }

    # ═══ PR ROW (row 10) — éditable, aligné sous chaque colonne produit ═══
    ws.cell(row=10, column=COL_LBL, value="Prix de revient (HT)").font = FONT_LABEL_MUT
    ws.cell(row=10, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)

    pr_row_data = [
        (COL_C_PRIX, COL_C_FLAG, 21.00, DOT_CONC),
        (COL_I_PRIX, COL_I_FLAG, 25.00, DOT_ISO),
        (COL_R_PRIX, COL_R_FLAG, 12.15, DOT_REC),
    ]
    for start_col, end_col, value, color in pr_row_data:
        ws.merge_cells(start_row=10, start_column=start_col, end_row=10, end_column=end_col)
        c = ws.cell(row=10, column=start_col, value=value)
        c.number_format = NUM_EUR
        c.font = Font(name="Consolas", size=10, bold=True, color=color)
        c.fill = FILL_AMBER
        c.alignment = Alignment(horizontal="right", indent=1)
        c.border = BOX
    ws.row_dimensions[10].height = 18

    # ═══ DÉCOMPOSITION P&L (rows 12-19) ═══
    write_section_header(ws, 11, "DÉCOMPOSITION P&L")

    # Row 12 : Prix Public TTC
    write_pnl_row(ws, 12, "Prix Public TTC",
                  lambda ret, pr: f"={ret}",
                  retenus)

    # Row 13 : - TVA
    write_pnl_row(ws, 13, "  (−) TVA 5,5 %",
                  lambda ret, pr: f"=-{ret}*{tva_cell}/(1+{tva_cell})",
                  retenus, is_negative=True)

    # Row 14 : Prix Public HT (sous-total italique)
    write_pnl_row(ws, 14, "Prix Public HT",
                  lambda ret, pr: f"={ret}/(1+{tva_cell})",
                  retenus, italic=True, top_border=True)

    # Row 15 : - Remise
    write_pnl_row(ws, 15, "  (−) Remise client 18 %",
                  lambda ret, pr: f"=-({ret}/(1+{tva_cell}))*{remise_cell}",
                  retenus, is_negative=True)

    # Row 16 : PVM net HT (sous-total italique)
    write_pnl_row(ws, 16, "PVM net HT",
                  lambda ret, pr: f"=({ret}/(1+{tva_cell}))*(1-{remise_cell})",
                  retenus, italic=True, top_border=True)

    # Row 17 : - PR
    write_pnl_row(ws, 17, "  (−) Prix de revient",
                  lambda ret, pr: f"=-{pr}",
                  retenus, is_negative=True)

    # Row 18 : Marge nette € (total gras top border)
    write_pnl_row(ws, 18, "Marge nette €",
                  lambda ret, pr: f"=({ret}/(1+{tva_cell}))*(1-{remise_cell})-{pr}",
                  retenus, bold=True, top_border=True)

    # Row 19 : % Marge nette (total gras)
    write_pnl_row(ws, 19, "% Marge nette",
                  lambda ret, pr: f"=IFERROR((({ret}/(1+{tva_cell}))*(1-{remise_cell})-{pr})/(({ret}/(1+{tva_cell}))*(1-{remise_cell})),0)",
                  retenus, bold=True, is_pct=True)

    # Row 20 : blank

    # Row 21 : Recommandation (flag formulas)
    ws.cell(row=21, column=COL_LBL, value="Recommandation").font = FONT_LABEL_B
    ws.cell(row=21, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)

    def flag_formula_product(retenu_cell, pr_cell, concentrate_ref=None, cannib_threshold=0.10):
        marge_expr = (
            f"(({retenu_cell}/(1+{tva_cell}))*(1-{remise_cell})-{pr_cell})"
            f"/(({retenu_cell}/(1+{tva_cell}))*(1-{remise_cell}))"
        )
        base = (
            f'IF({marge_expr}<{floor_cell},"⚠️ Marge basse",'
            f'IF({marge_expr}<{target_cell},"🟡 Limite",'
            f'IF({marge_expr}<={ceil_cell},"🎯 Optimal","🔵 Premium")))'
        )
        if concentrate_ref:
            return (
                f'=IF(({retenu_cell}-{concentrate_ref})/{concentrate_ref}<{cannib_threshold},"🚫 Cannibalisme",'
                f'{base})'
            )
        return f'={base}'

    flag_products = [
        ("conc", COL_C_PRIX, COL_C_FLAG, flag_formula_product(conc_retenu, pr_conc)),
        ("iso",  COL_I_PRIX, COL_I_FLAG, flag_formula_product(iso_retenu, pr_iso, conc_retenu, 0.10)),
        ("rec",  COL_R_PRIX, COL_R_FLAG, flag_formula_product(rec_retenu, pr_rec, conc_retenu, 0.05)),
    ]
    for key, start, end, formula in flag_products:
        ws.merge_cells(start_row=21, start_column=start, end_row=21, end_column=end)
        c = ws.cell(row=21, column=start, value=formula)
        c.font = FONT_FLAG_SIZE
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[21].height = 20

    # ═══ BENCHMARK MARCHÉ (rows 23-27) ═══
    write_section_header(ws, 23, "BENCHMARK MARCHÉ  (prix concurrents NET après code client standard)")

    bench_refs = {
        "conc": CONC_BENCH,
        "iso":  ISO_BENCH,
        "rec":  REC_BENCH,
    }

    bench_rows = [
        (24, "Médiane concurrence",   lambda refs: f"=MEDIAN({','.join(refs)})"),
        (25, "Moyenne concurrence",    lambda refs: f"=AVERAGE({','.join(refs)})"),
        (26, "Min / Max",              None),  # special case, two values
    ]

    for row, label, formula_fn in bench_rows[:2]:
        ws.cell(row=row, column=COL_LBL, value=label).font = FONT_LABEL
        ws.cell(row=row, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)
        for key, start, end in [
            ("conc", COL_C_PRIX, COL_C_FLAG),
            ("iso",  COL_I_PRIX, COL_I_FLAG),
            ("rec",  COL_R_PRIX, COL_R_FLAG),
        ]:
            ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
            c = ws.cell(row=row, column=start, value=formula_fn(bench_refs[key]))
            c.number_format = NUM_EUR
            c.font = FONT_NUM
            c.alignment = Alignment(horizontal="right", indent=1)

    # Row 26 : Min / Max as "26,00 € — 45,00 €" style
    ws.cell(row=26, column=COL_LBL, value="Fourchette min-max").font = FONT_LABEL
    ws.cell(row=26, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)
    for key, start, end in [
        ("conc", COL_C_PRIX, COL_C_FLAG),
        ("iso",  COL_I_PRIX, COL_I_FLAG),
        ("rec",  COL_R_PRIX, COL_R_FLAG),
    ]:
        refs = bench_refs[key]
        bl = ",".join(refs)
        ws.merge_cells(start_row=26, start_column=start, end_row=26, end_column=end)
        c = ws.cell(row=26, column=start, value=f'=TEXT(MIN({bl}),"0") & " €  —  " & TEXT(MAX({bl}),"0") & " €"')
        c.font = FONT_NUM_MUTED
        c.alignment = Alignment(horizontal="right", indent=1)

    # Row 27 : Écart retenu vs médiane
    ws.cell(row=27, column=COL_LBL, value="Écart retenu vs médiane").font = FONT_LABEL
    ws.cell(row=27, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)
    for key, start, end, retenu in [
        ("conc", COL_C_PRIX, COL_C_FLAG, conc_retenu),
        ("iso",  COL_I_PRIX, COL_I_FLAG, iso_retenu),
        ("rec",  COL_R_PRIX, COL_R_FLAG, rec_retenu),
    ]:
        refs = bench_refs[key]
        med = f"MEDIAN({','.join(refs)})"
        ws.merge_cells(start_row=27, start_column=start, end_row=27, end_column=end)
        c = ws.cell(row=27, column=start, value=f"=IFERROR(({retenu}-{med})/{med},0)")
        c.number_format = NUM_PCT_SIGNED
        c.font = FONT_NUM
        c.alignment = Alignment(horizontal="right", indent=1)

    # ═══ ALGO HYBRIDE MARCHÉ/MARGE (rows 29-36) ═══
    write_section_header(ws, 29, "ALGO PP OPTIMAL  (hybride marché-first)")

    # Build algo rows
    def floor_margin_formula(pr_cell, target):
        """Price TTC for a target HT margin, rounded up to X.9."""
        return f"CEILING({pr_cell}/(1-{target})/(1-{remise_cell})*(1+{tva_cell})+0.1,1)-0.1"

    pr_by_product = {"conc": pr_conc, "iso": pr_iso, "rec": pr_rec}

    algo_rows = [
        (30, "Plancher marge (⚠️)",   lambda k: f"={floor_margin_formula(pr_by_product[k], floor_cell)}"),
        (31, "Cible marge (🎯)",       lambda k: f"={floor_margin_formula(pr_by_product[k], target_cell)}"),
        (32, "Plafond marge (🔵)",     lambda k: f"={floor_margin_formula(pr_by_product[k], ceil_cell)}"),
        (33, "Médiane marché",         lambda k: f"=MEDIAN({','.join(bench_refs[k])})"),
        (34, "Anti-cannib",            None),  # special
        (35, "→ PP OPTIMAL algo",      None),  # special
    ]

    for row, label, formula_fn in algo_rows:
        ws.cell(row=row, column=COL_LBL, value=label).font = FONT_LABEL
        ws.cell(row=row, column=COL_LBL).alignment = Alignment(horizontal="left", indent=2)

        if row == 34:
            # Anti-cannib : "—" for Concentrate, Concentrate×1.10 for Iso, Concentrate×1.05 for Rec
            for key, start, end, formula in [
                ("conc", COL_C_PRIX, COL_C_FLAG, '="—"'),
                ("iso",  COL_I_PRIX, COL_I_FLAG, f"={floor_margin_formula('(' + conc_retenu + '*1.10/(1+' + tva_cell + ')*(1+' + tva_cell + '))', 0)}" if False else f"=CEILING({conc_retenu}*1.10+0.1,1)-0.1"),
                ("rec",  COL_R_PRIX, COL_R_FLAG, f"=CEILING({conc_retenu}*1.05+0.1,1)-0.1"),
            ]:
                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
                c = ws.cell(row=row, column=start, value=formula)
                if key == "conc":
                    c.font = FONT_NUM_MUTED
                    c.alignment = Alignment(horizontal="center")
                else:
                    c.number_format = NUM_EUR
                    c.font = FONT_NUM
                    c.alignment = Alignment(horizontal="right", indent=1)
        elif row == 35:
            # PP OPTIMAL hybride : marché si marge à médiane ≥ 30%, sinon plancher 30%
            # Then clamp up to anti-cannib for Iso/Rec
            for key, start, end, retenu_cell, bench_list, anti_cannib_formula in [
                ("conc", COL_C_PRIX, COL_C_FLAG, conc_retenu, ",".join(CONC_BENCH),
                 None),
                ("iso",  COL_I_PRIX, COL_I_FLAG, iso_retenu, ",".join(ISO_BENCH),
                 f"CEILING({conc_retenu}*1.10+0.1,1)-0.1"),
                ("rec",  COL_R_PRIX, COL_R_FLAG, rec_retenu, ",".join(REC_BENCH),
                 f"CEILING({conc_retenu}*1.05+0.1,1)-0.1"),
            ]:
                pr_cell = pr_by_product[key]
                plancher = floor_margin_formula(pr_cell, floor_cell)
                med = f"MEDIAN({bench_list})"
                marge_at_med = (
                    f"((({med}/(1+{tva_cell}))*(1-{remise_cell})-{pr_cell})"
                    f"/(({med}/(1+{tva_cell}))*(1-{remise_cell})))"
                )
                hybride = f"IF({marge_at_med}<{floor_cell},{plancher},{med})"
                if anti_cannib_formula:
                    final = f"=MAX({hybride},{anti_cannib_formula})"
                else:
                    final = f"={hybride}"

                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
                c = ws.cell(row=row, column=start, value=final)
                c.number_format = NUM_EUR
                c.font = FONT_NUM_B
                c.alignment = Alignment(horizontal="right", indent=1)
                c.border = TOP_INK
        else:
            for key, start, end in [
                ("conc", COL_C_PRIX, COL_C_FLAG),
                ("iso",  COL_I_PRIX, COL_I_FLAG),
                ("rec",  COL_R_PRIX, COL_R_FLAG),
            ]:
                ws.merge_cells(start_row=row, start_column=start, end_row=row, end_column=end)
                c = ws.cell(row=row, column=start, value=formula_fn(key))
                c.number_format = NUM_EUR
                c.font = FONT_NUM_MUTED
                c.alignment = Alignment(horizontal="right", indent=1)

    # ═══ OPTIONS 1-10 (rows 37-51) ═══
    # Option 1 (row 42) = pivot ÉDITABLE (cellule ambre) par produit.
    # Options 2-10 cascadent en formule +1/+2/+3/+4/+5/+6/+8/+10/+12 € depuis le pivot.
    write_section_header(ws, 37, "OPTIONS DE PRIX  ·  Option 1 = PIVOT ÉDITABLE (ambre)  ·  cascade +1 à +12 €")

    # Row 38 : Sub-headers
    ws.cell(row=38, column=COL_LBL, value="#").font = FONT_SECTION
    ws.cell(row=38, column=COL_LBL).alignment = Alignment(horizontal="center")

    sub_headers = [
        (COL_C_PRIX, "Prix"),
        (COL_C_MARGE, "Marge"),
        (COL_C_FLAG, "Flag"),
        (COL_I_PRIX, "Prix"),
        (COL_I_MARGE, "Marge"),
        (COL_I_FLAG, "Flag"),
        (COL_R_PRIX, "Prix"),
        (COL_R_MARGE, "Marge"),
        (COL_R_FLAG, "Flag"),
    ]
    for col, label in sub_headers:
        c = ws.cell(row=38, column=col, value=label)
        c.font = FONT_SECTION
        c.alignment = Alignment(horizontal="center")
    # Bottom border under sub-headers
    for col in range(1, COL_DELTA + 1):
        ws.cell(row=38, column=col).border = BOT_INK

    ws.row_dimensions[38].height = 16

    # Rows 39-41 : blank (reserve)
    # Options start at row options_start_row (42)

    def margin_formula(prix_cell, pr_cell):
        pvm = f"(({prix_cell}/(1+{tva_cell}))*(1-{remise_cell}))"
        return f"=IFERROR(({pvm}-{pr_cell})/{pvm},0)"

    def option_flag(marge_cell, prix_cell=None, concentrate_ref=None, threshold=0.10):
        base = (
            f'IF({marge_cell}<{floor_cell},"⚠️",'
            f'IF({marge_cell}<{target_cell},"🟡",'
            f'IF({marge_cell}<={ceil_cell},"🎯","🔵")))'
        )
        if concentrate_ref and prix_cell:
            cannib_expr = f"({prix_cell}-{concentrate_ref})/{concentrate_ref}"
            return f'=IF(IFERROR({cannib_expr},0)<{threshold},"🚫",{base})'
        return f'={base}'

    # Pivot cells (option 1, row 42) = éditables par produit
    pivot_conc_cell = f"${get_column_letter(COL_C_PRIX)}${options_start_row}"
    pivot_iso_cell  = f"${get_column_letter(COL_I_PRIX)}${options_start_row}"
    pivot_rec_cell  = f"${get_column_letter(COL_R_PRIX)}${options_start_row}"

    pivots = [
        (COL_C_PRIX, pivot_conc_cell, pr_conc, None,         None),
        (COL_I_PRIX, pivot_iso_cell,  pr_iso,  conc_retenu,  0.10),
        (COL_R_PRIX, pivot_rec_cell,  pr_rec,  conc_retenu,  0.05),
    ]

    for i in range(10):
        r = options_start_row + i
        opt_num = i + 1
        offset = OFFSETS[i]

        # Option number (col A)
        c = ws.cell(row=r, column=COL_LBL, value=opt_num)
        c.font = Font(name="Arial", size=9, color=MUTED)
        c.alignment = Alignment(horizontal="center")

        for col_prix, pivot_cell, pr_cell, cannib_ref, cannib_thr in pivots:
            col_marge = col_prix + 1
            col_flag  = col_prix + 2

            prix_cell_ref = f"{get_column_letter(col_prix)}{r}"
            marge_cell_ref = f"{get_column_letter(col_marge)}{r}"

            # Row 42 (option 1) : pivot ÉDITABLE (valeur en dur, ambre + border)
            # Rows 43-51 (option 2-10) : formule cascade depuis pivot
            if i == 0:
                c_prix = ws.cell(row=r, column=col_prix, value=PIVOT_DEFAULT)
                c_prix.fill = FILL_AMBER
                c_prix.border = BOX
                c_prix.font = Font(name="Consolas", size=10, bold=True, color=INK)
            else:
                c_prix = ws.cell(
                    row=r, column=col_prix,
                    value=f"=CEILING({pivot_cell}+{offset}+0.1,1)-0.1",
                )
                c_prix.font = FONT_NUM
            c_prix.number_format = NUM_EUR
            c_prix.alignment = Alignment(horizontal="right", indent=1)

            # Marge
            c_marge = ws.cell(row=r, column=col_marge, value=margin_formula(prix_cell_ref, pr_cell))
            c_marge.number_format = NUM_PCT
            c_marge.font = FONT_NUM
            c_marge.alignment = Alignment(horizontal="right", indent=1)

            # Flag (anti-cannib uniquement pour Iso/Rec)
            if cannib_ref:
                flag_val = option_flag(marge_cell_ref, prix_cell_ref, cannib_ref, cannib_thr)
            else:
                flag_val = option_flag(marge_cell_ref)
            c_flag = ws.cell(row=r, column=col_flag, value=flag_val)
            c_flag.alignment = Alignment(horizontal="center")

        # Zebra striping (skip if row 42 already has amber pivot fills)
        if i % 2 == 1:
            for col in range(1, COL_DELTA + 1):
                cell = ws.cell(row=r, column=col)
                if not cell.fill.start_color or cell.fill.start_color.value in (None, "00000000"):
                    cell.fill = FILL_LIGHT

        ws.row_dimensions[r].height = 15

    # ─── Conditional formatting : highlight la row sélectionnée par produit ───
    # Fill ambre medium + bold sur le cells BCD (Conc), FGH (Iso), JKL (Rec)
    # en fonction de la dropdown du produit.
    # Col A (#) highlight si n'importe quel produit a sélectionné cette row.
    fill_selected = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")
    font_selected = Font(name="Consolas", size=10, bold=True, color=INK)

    rule_conc = FormulaRule(
        formula=[f"$A{options_start_row}=$B$8"],
        fill=fill_selected,
        font=font_selected,
    )
    ws.conditional_formatting.add(
        f"B{options_start_row}:D{options_end_row}",
        rule_conc,
    )

    rule_iso = FormulaRule(
        formula=[f"$A{options_start_row}=$F$8"],
        fill=fill_selected,
        font=font_selected,
    )
    ws.conditional_formatting.add(
        f"F{options_start_row}:H{options_end_row}",
        rule_iso,
    )

    rule_rec = FormulaRule(
        formula=[f"$A{options_start_row}=$J$8"],
        fill=fill_selected,
        font=font_selected,
    )
    ws.conditional_formatting.add(
        f"J{options_start_row}:L{options_end_row}",
        rule_rec,
    )

    # Col A (# col) — highlight if ANY product has selected this row
    rule_hash = FormulaRule(
        formula=[f"OR($A{options_start_row}=$B$8,$A{options_start_row}=$F$8,$A{options_start_row}=$J$8)"],
        font=Font(name="Arial", size=9, bold=True, color=INK),
    )
    ws.conditional_formatting.add(
        f"A{options_start_row}:A{options_end_row}",
        rule_hash,
    )

    # ─── Column widths ───
    widths = {
        "A": 28,  # labels
        "B": 10, "C": 9, "D": 13,   # Concentrate
        "E": 1,                      # gap (narrow)
        "F": 10, "G": 9, "H": 13,   # Isolate
        "I": 1,                      # gap
        "J": 10, "K": 9, "L": 13,   # Recovery
        "M": 1,                      # right margin
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A7"


# ═════════════════════════════════════════════════════════════════════════
def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    # Clean up old pricing tabs
    for old_name in (
        "Pricing Whey Concentrate", "Pricing Whey Isolate", "Pricing Whey Recovery",
        "Pricing Dynamique", "Pricing Grille C", "Pricing Dashboard A",
        "Pricing Grille",
    ):
        if old_name in wb.sheetnames:
            del wb[old_name]

    build_pricing_v10(wb)

    wb.save(DST)
    print(f"✓ Generated {DST}")
    print(f"  Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
