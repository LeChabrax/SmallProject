"""
Build v9 — HT-based margin + dynamic PR + P&L vertical section.

Changes vs v8 :
  1. Margin formulas switched from TTC-based to HT-based :
     PP HT     = PP TTC / (1 + TVA)
     PVM net   = PP HT × (1 − remise)
     Marge €   = PVM net − PR
     % Marge   = Marge € / PVM net
  2. PP OPTIMAL auto-calc now uses HT-based math then converts back to TTC :
     Plancher 38% = CEILING(PR/(1-0.38)/(1-remise) × (1+TVA) + 0.1, 1) − 0.1
     Cible 40%    = CEILING(PR/(1-0.40)/(1-remise) × (1+TVA) + 0.1, 1) − 0.1
  3. Added P&L vertical section (rows 40-49) below the grid with 8 lines :
     Prix TTC / -TVA / =PP HT / -Remise / =PVM net / -PR / =Marge € / % Marge
  4. Reduced font sizes : prix retenu 20pt → 14pt, marge 14pt → 11pt
  5. Removed Pricing Dashboard A tab (user prefers Grille C only)
  6. PR cells in Hypothèses stay editable — changing them cascades everything
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path("/Users/antoinechabrat/Downloads/Copie de benchmark_whey_concurrent 5 1.xlsx")
DST = Path(
    "/Users/antoinechabrat/Documents/SmallProject/Impulse-Nutrition/benchmark/"
    "concurrent/benchmark_whey_concurrent_v9_dynamic.xlsx"
)

# ─── Design tokens (slightly smaller fonts than v8) ──────────────────────
INK       = "0F172A"
MUTED     = "64748B"
BORDER    = "CBD5E1"
LIGHT     = "F1F5F9"
WHITE     = "FFFFFF"
ACCENT    = "DC2626"
AMBER     = "FEF3C7"
AMBER_STR = "F59E0B"

COL_CONC  = "1D4ED8"  # blue-700
COL_ISO   = "047857"  # emerald-700
COL_REC   = "B91C1C"  # red-700

FONT_TITLE     = Font(name="Arial", size=16, bold=True, color=INK)
FONT_SUBTITLE  = Font(name="Arial", size=9, italic=True, color=MUTED)
FONT_SECTION   = Font(name="Arial", size=11, bold=True, color=WHITE)
FONT_LABEL     = Font(name="Arial", size=9, bold=True, color=INK)
FONT_LABEL_MUT = Font(name="Arial", size=8, italic=True, color=MUTED)
FONT_HEADER    = Font(name="Arial", size=9, bold=True, color=WHITE)
FONT_NUM       = Font(name="Consolas", size=9, color=INK)
FONT_NUM_BOLD  = Font(name="Consolas", size=10, bold=True, color=INK)
FONT_PRIX_RETENU  = Font(name="Arial", size=14, bold=True, color=ACCENT)      # était 20pt
FONT_MARGE_RETENU = Font(name="Arial", size=11, bold=True, color=INK)         # était 14pt
FONT_FLAG      = Font(name="Arial", size=9, color=INK)
FONT_DROPDOWN  = Font(name="Arial", size=12, bold=True, color=INK)
FONT_PNL_TOT   = Font(name="Consolas", size=11, bold=True, color=INK)

FILL_CONC      = PatternFill(start_color=COL_CONC, end_color=COL_CONC, fill_type="solid")
FILL_ISO       = PatternFill(start_color=COL_ISO, end_color=COL_ISO, fill_type="solid")
FILL_REC       = PatternFill(start_color=COL_REC, end_color=COL_REC, fill_type="solid")
FILL_MUTED     = PatternFill(start_color=MUTED, end_color=MUTED, fill_type="solid")
FILL_LIGHT     = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
FILL_AMBER     = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
FILL_AMBER_STR = PatternFill(start_color=AMBER_STR, end_color=AMBER_STR, fill_type="solid")
FILL_ACCENT    = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
FILL_INK       = PatternFill(start_color=INK, end_color=INK, fill_type="solid")

THIN_BORDER = Side(style="thin", color=BORDER)
BOX_BORDER  = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)
TOP_LINE    = Border(top=Side(style="thin", color=INK))
BOTTOM_LINE = Border(bottom=Side(style="thin", color=INK))

NUM_EUR         = '#,##0.00" €"'
NUM_EUR_NEG     = '#,##0.00" €";[Red]-#,##0.00" €"'
NUM_PCT         = "0.0%"
NUM_PCT_SIGNED  = '+0.0%;-0.0%;0.0%'

# Benchmark refs (same as v8 — don't touch)
CONC_BENCH = [
    "'Synthèse prix'!H14", "'Synthèse prix'!H15", "'Synthèse prix'!H19",
    "'Synthèse prix'!H20", "'Synthèse prix'!H24", "'Synthèse prix'!H25",
    "'Synthèse prix'!H26", "'Synthèse prix'!H31", "'Synthèse prix'!H37",
    "'Synthèse prix'!H50",
]
ISO_BENCH = [
    "'Synthèse prix'!H7", "'Synthèse prix'!H10", "'Synthèse prix'!H11",
    "'Synthèse prix'!H18", "'Synthèse prix'!H21", "'Synthèse prix'!H22",
    "'Synthèse prix'!H23", "'Synthèse prix'!H29", "'Synthèse prix'!H30",
    "'Synthèse prix'!H35", "'Synthèse prix'!H41", "'Synthèse prix'!H47",
]
REC_BENCH = [
    "'Synthèse prix'!H5", "'Synthèse prix'!H6", "'Synthèse prix'!H32",
]

OFFSETS = [-4, -3, -2, -1, 0, 1, 2, 3, 5, 7]
CONC_PRICES = [29.9, 31.9, 33.9, 34.9, 36.9, 37.9, 39.9, 41.9, 42.9, 44.9]


def add_dropdown(ws, cell_ref):
    dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_ref)
    ws[cell_ref] = 5
    ws[cell_ref].font = FONT_DROPDOWN
    ws[cell_ref].fill = FILL_AMBER_STR
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell_ref].border = BOX_BORDER


def flag_formula_concentrate(marge_cell):
    return (
        f'=IF({marge_cell}<0.30,"⚠️ Marge basse",'
        f'IF({marge_cell}<0.35,"🟡 Limite",'
        f'IF({marge_cell}<=0.42,"🎯 Optimal","🔵 Premium")))'
    )


def flag_formula_cascading(marge_cell, prix_cell, concentrate_ref, cannib_threshold=0.10):
    return (
        f'=IF({marge_cell}<0.30,"⚠️ Marge basse",'
        f'IF(IFERROR(({prix_cell}-{concentrate_ref})/{concentrate_ref},0)<{cannib_threshold},"🚫 Cannibalisme",'
        f'IF({marge_cell}<0.35,"🟡 Limite",'
        f'IF({marge_cell}<=0.42,"🎯 Optimal","🔵 Premium"))))'
    )


def ht_margin_formula(prix_cell, tva_cell, remise_cell, pr_cell):
    """Return the HT-based margin percentage formula as string.

    % Marge = ((PP_TTC / (1+TVA)) × (1 − remise) − PR) / ((PP_TTC / (1+TVA)) × (1 − remise))
    """
    pvm = f"(({prix_cell}/(1+{tva_cell}))*(1-{remise_cell}))"
    return f"=IFERROR(({pvm}-{pr_cell})/{pvm},0)"


def ht_optimal_formula(pr_cell, target_margin, remise_cell, tva_cell):
    """Return a PP TTC formula that gives exactly target_margin HT, rounded up to X.9.

    From PR / (1−M) → PVM net HT → / (1−remise) → × (1+TVA) → PP TTC
    """
    return (
        f"=CEILING({pr_cell}/(1-{target_margin})/(1-{remise_cell})*(1+{tva_cell})+0.1,1)-0.1"
    )


# ═════════════════════════════════════════════════════════════════════════
#  BUILD PRICING GRILLE (v9) — single tab with 3 stacked sections + P&L
# ═════════════════════════════════════════════════════════════════════════
def build_pricing_grille(wb):
    if "Pricing Grille" in wb.sheetnames:
        del wb["Pricing Grille"]
    ws = wb.create_sheet("Pricing Grille", 0)

    # ── Row 1 : Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "PRICING DYNAMIQUE · GAMME WHEY IMPULSE"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30

    # ── Row 2 : Subtitle
    ws.merge_cells("A2:N2")
    ws["A2"] = "HT-based · 3 produits en cascade · modifier un PR ou un dropdown recalcule tout"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Row 4 : Hypothèses label
    ws["A4"] = "HYPOTHÈSES PARTAGÉES  ·  cellules éditables"
    ws["A4"].font = Font(name="Arial", size=8, bold=True, color=MUTED)
    ws.merge_cells("A4:N4")

    # ── Row 5 : Inline hypothèses — EDITABLE cells
    hyp = [
        (1, "TVA", 0.055, NUM_PCT),
        (3, "Remise", 0.18, NUM_PCT),
        (5, "PR Conc", 21, NUM_EUR),
        (7, "PR Iso", 25, NUM_EUR),
        (9, "PR Rec", 12.15, NUM_EUR),
    ]
    for col, label, value, fmt in hyp:
        c = ws.cell(row=5, column=col, value=label)
        c.font = FONT_LABEL_MUT
        c.alignment = Alignment(horizontal="right")
        v = ws.cell(row=5, column=col + 1, value=value)
        v.number_format = fmt
        v.font = FONT_NUM_BOLD
        v.fill = FILL_AMBER  # signals "editable"
        v.alignment = Alignment(horizontal="center")
        v.border = BOX_BORDER

    tva_cell    = "$B$5"
    remise_cell = "$D$5"
    pr_conc     = "$F$5"
    pr_iso      = "$H$5"
    pr_rec      = "$J$5"

    # ── Row 7 : Section headers (3 bandeaux)
    ws.merge_cells("A7:D7")
    ws["A7"] = "▌ CONCENTRATE"
    ws["A7"].font = FONT_SECTION
    ws["A7"].fill = FILL_CONC
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 22

    ws.merge_cells("F7:I7")
    ws["F7"] = "▌ ISOLATE   ⬅ cascade depuis Concentrate retenu"
    ws["F7"].font = FONT_SECTION
    ws["F7"].fill = FILL_ISO
    ws["F7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("K7:N7")
    ws["K7"] = "▌ RECOVERY   ⬅ cascade depuis C + I"
    ws["K7"].font = FONT_SECTION
    ws["K7"].fill = FILL_REC
    ws["K7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Row 9 : Option label
    for col in [1, 6, 11]:
        c = ws.cell(row=9, column=col, value="🎯 Option sélectionnée")
        c.font = FONT_LABEL_MUT
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 3)

    # ── Row 10 : DROPDOWNS
    conc_dropdown = "B10"
    iso_dropdown  = "G10"
    rec_dropdown  = "L10"
    for cell_ref in [conc_dropdown, iso_dropdown, rec_dropdown]:
        add_dropdown(ws, cell_ref)
    ws.row_dimensions[10].height = 22
    ws.merge_cells("B10:D10")
    ws.merge_cells("G10:I10")
    ws.merge_cells("L10:N10")

    # ── Row 12 : Prix retenu label
    for col in [1, 6, 11]:
        c = ws.cell(row=12, column=col, value="→ Prix retenu")
        c.font = FONT_LABEL_MUT
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=12, start_column=col, end_row=12, end_column=col + 3)

    # ── Row 13 : BIG prix retenu (14pt)
    options_start_row = 29
    options_end_row = 38

    conc_retenu_cell = "B13"
    iso_retenu_cell  = "G13"
    rec_retenu_cell  = "L13"

    ws[conc_retenu_cell] = f"=INDEX($B${options_start_row}:$B${options_end_row},${conc_dropdown})"
    ws[iso_retenu_cell]  = f"=INDEX($G${options_start_row}:$G${options_end_row},${iso_dropdown})"
    ws[rec_retenu_cell]  = f"=INDEX($L${options_start_row}:$L${options_end_row},${rec_dropdown})"

    for cell_ref in [conc_retenu_cell, iso_retenu_cell, rec_retenu_cell]:
        c = ws[cell_ref]
        c.font = FONT_PRIX_RETENU
        c.fill = FILL_AMBER
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = NUM_EUR
        c.border = BOX_BORDER
    ws.merge_cells("B13:D13")
    ws.merge_cells("G13:I13")
    ws.merge_cells("L13:N13")
    ws.row_dimensions[13].height = 22

    # ── Row 14 : Marge retenue (HT-based formulas)
    for label_col, pr_cell, retenu_cell, merge_end in [
        ("A", pr_conc, conc_retenu_cell, "D"),
        ("F", pr_iso, iso_retenu_cell, "I"),
        ("K", pr_rec, rec_retenu_cell, "N"),
    ]:
        lbl_col_letter = label_col
        val_col_letter = chr(ord(label_col) + 1)
        ws[f"{lbl_col_letter}14"] = "Marge nette HT"
        ws[f"{lbl_col_letter}14"].font = FONT_LABEL_MUT
        ws[f"{lbl_col_letter}14"].alignment = Alignment(horizontal="left", indent=1)
        # Marge formula
        ws[f"{val_col_letter}14"] = ht_margin_formula(retenu_cell, tva_cell, remise_cell, pr_cell)
        ws[f"{val_col_letter}14"].font = FONT_MARGE_RETENU
        ws[f"{val_col_letter}14"].number_format = NUM_PCT
        ws[f"{val_col_letter}14"].alignment = Alignment(horizontal="center", vertical="center")
        # Merge value across remaining cols
        from_col_idx = ord(val_col_letter) - 64
        to_col_idx = ord(merge_end) - 64
        ws.merge_cells(start_row=14, start_column=from_col_idx, end_row=14, end_column=to_col_idx)
    ws.row_dimensions[14].height = 18

    # ── Row 16 : BENCHMARK label
    for col in [1, 6, 11]:
        c = ws.cell(row=16, column=col, value="BENCHMARK MARCHÉ")
        c.font = Font(name="Arial", size=8, bold=True, color=MUTED)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=16, start_column=col, end_row=16, end_column=col + 3)

    # ── Row 17 : Stats values + Row 18 : sub-labels
    for col, bench in [(1, CONC_BENCH), (6, ISO_BENCH), (11, REC_BENCH)]:
        bench_list = ",".join(bench)
        stats = [
            ("Moy", f"=AVERAGE({bench_list})"),
            ("Méd", f"=MEDIAN({bench_list})"),
            ("Min", f"=MIN({bench_list})"),
            ("Max", f"=MAX({bench_list})"),
        ]
        for i, (label, formula) in enumerate(stats):
            v = ws.cell(row=17, column=col + i, value=formula)
            v.number_format = NUM_EUR
            v.font = FONT_NUM
            v.alignment = Alignment(horizontal="center")
            lbl = ws.cell(row=18, column=col + i, value=label)
            lbl.font = Font(name="Arial", size=8, italic=True, color=MUTED)
            lbl.alignment = Alignment(horizontal="center")

    # Median cells (col 3, 8, 13 = C, H, M)
    conc_median = "C17"
    iso_median  = "H17"
    rec_median  = "M17"

    # ── Row 20 : CALCUL AUTO header (Isolate + Recovery)
    for col in [6, 11]:
        c = ws.cell(row=20, column=col, value="CALCUL AUTO — Algo Option C (HT-based)")
        c.font = Font(name="Arial", size=8, bold=True, color=MUTED)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=20, start_column=col, end_row=20, end_column=col + 3)

    # ── Rows 21-24 : 4 calc rows per product (Isolate + Recovery only)
    iso_bench_list = ",".join(ISO_BENCH)
    rec_bench_list = ",".join(REC_BENCH)

    # Isolate calcul auto
    iso_calc = [
        ("Plancher marge 38%",  ht_optimal_formula(pr_iso, 0.38, remise_cell, tva_cell)),
        ("Cible marge 40%",     ht_optimal_formula(pr_iso, 0.40, remise_cell, tva_cell)),
        ("Anti-cannib (C×1,10)", f"=CEILING({conc_retenu_cell}*1.10+0.1,1)-0.1"),
        ("Marché médiane",      f"=MEDIAN({iso_bench_list})"),
    ]
    for i, (label, formula) in enumerate(iso_calc):
        r = 21 + i
        lbl = ws.cell(row=r, column=6, value=label)
        lbl.font = FONT_LABEL_MUT
        lbl.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        v = ws.cell(row=r, column=9, value=formula)
        v.font = FONT_NUM
        v.number_format = NUM_EUR
        v.alignment = Alignment(horizontal="center")

    iso_marge38 = "I21"
    iso_marge40 = "I22"
    iso_anticannib = "I23"
    iso_marche = "I24"
    iso_pp_opt = "I25"

    # Row 25 : PP OPTIMAL Isolate
    ws["F25"] = "→ PP OPTIMAL"
    ws["F25"].font = Font(name="Arial", size=9, bold=True, color=WHITE)
    ws["F25"].fill = FILL_ACCENT
    ws["F25"].alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("F25:H25")
    ws[iso_pp_opt] = f"=MAX({iso_marge38},{iso_anticannib},MIN({iso_marge40},{iso_marche}))"
    ws[iso_pp_opt].font = Font(name="Arial", size=11, bold=True, color=WHITE)
    ws[iso_pp_opt].fill = FILL_ACCENT
    ws[iso_pp_opt].number_format = NUM_EUR
    ws[iso_pp_opt].alignment = Alignment(horizontal="center", vertical="center")

    # Recovery calcul auto
    rec_calc = [
        ("Plancher marge 38%",  ht_optimal_formula(pr_rec, 0.38, remise_cell, tva_cell)),
        ("Plancher C+5%",       f"=CEILING({conc_retenu_cell}*1.05+0.1,1)-0.1"),
        ("Plafond I−5%",        f"=CEILING({iso_retenu_cell}*0.95+0.1,1)-0.1"),
        ("Marché réf (Impulse)", "='Synthèse prix'!H5"),
    ]
    for i, (label, formula) in enumerate(rec_calc):
        r = 21 + i
        lbl = ws.cell(row=r, column=11, value=label)
        lbl.font = FONT_LABEL_MUT
        lbl.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
        v = ws.cell(row=r, column=14, value=formula)
        v.font = FONT_NUM
        v.number_format = NUM_EUR
        v.alignment = Alignment(horizontal="center")

    rec_marge38 = "N21"
    rec_floor   = "N22"
    rec_ceiling = "N23"
    rec_marche  = "N24"
    rec_pp_opt  = "N25"

    ws["K25"] = "→ PP OPTIMAL"
    ws["K25"].font = Font(name="Arial", size=9, bold=True, color=WHITE)
    ws["K25"].fill = FILL_ACCENT
    ws["K25"].alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("K25:M25")
    ws[rec_pp_opt] = f"=MAX({rec_marge38},{rec_floor},MIN({rec_ceiling},{rec_marche}))"
    ws[rec_pp_opt].font = Font(name="Arial", size=11, bold=True, color=WHITE)
    ws[rec_pp_opt].fill = FILL_ACCENT
    ws[rec_pp_opt].number_format = NUM_EUR
    ws[rec_pp_opt].alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 27 : Table headers
    headers = ["#", "Prix", "Marge %", "Flag"]
    for group_start in [1, 6, 11]:
        for i, h in enumerate(headers):
            c = ws.cell(row=27, column=group_start + i, value=h)
            c.font = FONT_HEADER
            c.fill = FILL_MUTED
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[27].height = 18

    # ── Rows 29-38 : Options (HT-based margin)
    for i in range(10):
        r = options_start_row + i
        opt_num = i + 1

        # ─── Concentrate ───
        ws.cell(row=r, column=1, value=opt_num).font = Font(name="Arial", size=9, bold=True, color=MUTED)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=CONC_PRICES[i]).number_format = NUM_EUR
        ws.cell(row=r, column=3, value=ht_margin_formula(f"B{r}", tva_cell, remise_cell, pr_conc)).number_format = NUM_PCT
        ws.cell(row=r, column=4, value=flag_formula_concentrate(f"C{r}"))
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="left", indent=1)

        # ─── Isolate (dynamic from iso_pp_opt) ───
        offset = OFFSETS[i]
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        prix_expr_iso = f"{iso_pp_opt}" if offset == 0 else f"{iso_pp_opt}{sign}{abs_off}"

        ws.cell(row=r, column=6, value=opt_num).font = Font(name="Arial", size=9, bold=True, color=MUTED)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7, value=f"=CEILING({prix_expr_iso}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=r, column=8, value=ht_margin_formula(f"G{r}", tva_cell, remise_cell, pr_iso)).number_format = NUM_PCT
        ws.cell(row=r, column=9, value=flag_formula_cascading(f"H{r}", f"G{r}", conc_retenu_cell, 0.10))
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="left", indent=1)

        # ─── Recovery (dynamic from rec_pp_opt) ───
        prix_expr_rec = f"{rec_pp_opt}" if offset == 0 else f"{rec_pp_opt}{sign}{abs_off}"
        ws.cell(row=r, column=11, value=opt_num).font = Font(name="Arial", size=9, bold=True, color=MUTED)
        ws.cell(row=r, column=11).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=12, value=f"=CEILING({prix_expr_rec}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=r, column=13, value=ht_margin_formula(f"L{r}", tva_cell, remise_cell, pr_rec)).number_format = NUM_PCT
        ws.cell(row=r, column=14, value=flag_formula_cascading(f"M{r}", f"L{r}", conc_retenu_cell, 0.05))
        ws.cell(row=r, column=14).alignment = Alignment(horizontal="left", indent=1)

        # Fonts + alignment for numeric cells
        for col in [2, 3, 7, 8, 12, 13]:
            cell = ws.cell(row=r, column=col)
            cell.font = FONT_NUM
            cell.alignment = Alignment(horizontal="right", indent=1)

        # Zebra rows
        if i % 2 == 1:
            for col in [1, 2, 3, 4, 6, 7, 8, 9, 11, 12, 13, 14]:
                c = ws.cell(row=r, column=col)
                if not c.fill.start_color or c.fill.start_color.value in (None, "00000000"):
                    c.fill = FILL_LIGHT

        # Highlight center (#5)
        if opt_num == 5:
            for col in [1, 2, 3, 4, 6, 7, 8, 9, 11, 12, 13, 14]:
                ws.cell(row=r, column=col).fill = FILL_AMBER

        ws.row_dimensions[r].height = 16

    # ═════════════════════════════════════════════════════════════════════
    # ── P&L SECTION (rows 40-50)
    # ═════════════════════════════════════════════════════════════════════

    # Row 40 : Section header
    ws.merge_cells("A40:N40")
    ws["A40"] = "▌ DÉTAIL P&L — DÉCOMPOSITION DES PRIX RETENUS"
    ws["A40"].font = FONT_SECTION
    ws["A40"].fill = FILL_INK
    ws["A40"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[40].height = 22

    # Row 41 : Column headers for P&L
    pnl_headers = [(2, "CONCENTRATE", COL_CONC), (7, "ISOLATE", COL_ISO), (12, "RECOVERY", COL_REC)]
    for col, label, color in pnl_headers:
        c = ws.cell(row=41, column=col, value=label)
        c.font = Font(name="Arial", size=10, bold=True, color=color)
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=41, start_column=col, end_row=41, end_column=col + 2)

    # The 8 P&L lines, each with a label (col A) + 3 value cols (B/C, G/H, L/M)
    # Col A : label (full width)
    # Col B, G, L : value (right-aligned mono)
    # We use columns B-D for Concentrate (merged per value), G-I for Isolate, L-N for Recovery
    pnl_lines = [
        # (row, label, sign, formula_builder, show_as_negative, is_total)
        (42, "Prix Public TTC",      "",  lambda r: r, False, False),
        (43, "− TVA (5,5 %)",        "−", lambda r: f"={r}*{tva_cell}/(1+{tva_cell})", True, False),
        (44, "= Prix Public HT",     "=", lambda r: f"={r}/(1+{tva_cell})", False, True),
        (45, "− Remise client",      "−", lambda r: f"=({r}/(1+{tva_cell}))*{remise_cell}", True, False),
        (46, "= PVM net HT",         "=", lambda r: f"=({r}/(1+{tva_cell}))*(1-{remise_cell})", False, True),
        (47, "− Prix de revient",    "−", None, True, False),  # PR is a direct cell ref
        (48, "= Marge nette €",      "=", lambda r: f"=({r}/(1+{tva_cell}))*(1-{remise_cell})-{{PR}}", False, True),
        (49, "% Marge nette",         "", None, False, False),
    ]

    retenus = {
        "conc": (2, conc_retenu_cell, pr_conc),
        "iso":  (7, iso_retenu_cell, pr_iso),
        "rec":  (12, rec_retenu_cell, pr_rec),
    }

    for row, label, sign, formula_builder, is_neg, is_total in pnl_lines:
        # Column A : label
        c = ws.cell(row=row, column=1, value=label)
        c.font = FONT_LABEL if is_total else Font(name="Arial", size=9, color=INK)
        c.alignment = Alignment(horizontal="left", indent=1)

        for product, (start_col, retenu_cell, pr_cell) in retenus.items():
            v = ws.cell(row=row, column=start_col)

            if row == 42:  # Prix TTC = retenu direct
                v.value = f"={retenu_cell}"
            elif row == 43:  # − TVA
                v.value = f"=-{retenu_cell}*{tva_cell}/(1+{tva_cell})"
            elif row == 44:  # = PP HT
                v.value = f"={retenu_cell}/(1+{tva_cell})"
            elif row == 45:  # − Remise
                v.value = f"=-({retenu_cell}/(1+{tva_cell}))*{remise_cell}"
            elif row == 46:  # = PVM net HT
                v.value = f"=({retenu_cell}/(1+{tva_cell}))*(1-{remise_cell})"
            elif row == 47:  # − PR
                v.value = f"=-{pr_cell}"
            elif row == 48:  # = Marge nette €
                v.value = f"=({retenu_cell}/(1+{tva_cell}))*(1-{remise_cell})-{pr_cell}"
            elif row == 49:  # % Marge
                pvm = f"(({retenu_cell}/(1+{tva_cell}))*(1-{remise_cell}))"
                v.value = f"=IFERROR(({pvm}-{pr_cell})/{pvm},0)"

            if row == 49:
                v.number_format = NUM_PCT
                v.font = FONT_PNL_TOT
                v.alignment = Alignment(horizontal="center")
            elif is_total:
                v.number_format = NUM_EUR
                v.font = FONT_PNL_TOT
                v.alignment = Alignment(horizontal="right", indent=1)
                v.border = TOP_LINE
            else:
                v.number_format = NUM_EUR_NEG if is_neg else NUM_EUR
                v.font = FONT_NUM
                v.alignment = Alignment(horizontal="right", indent=1)

            # Merge value cell across 3 cols of the product group
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 2)

        # Row styling for total lines
        if is_total:
            ws.row_dimensions[row].height = 20
        elif row == 49:
            ws.row_dimensions[row].height = 24

    # ── Column widths
    widths = {
        "A": 24, "B": 10, "C": 11, "D": 16,
        "E": 2,
        "F": 5, "G": 10, "H": 11, "I": 16,
        "J": 2,
        "K": 5, "L": 10, "M": 11, "N": 16,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A7"


# ═════════════════════════════════════════════════════════════════════════
def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    # Clean up old pricing tabs
    for old_name in (
        "Pricing Whey Concentrate",
        "Pricing Whey Isolate",
        "Pricing Whey Recovery",
        "Pricing Dynamique",
        "Pricing Grille C",
        "Pricing Dashboard A",
        "Pricing Grille",
    ):
        if old_name in wb.sheetnames:
            del wb[old_name]

    build_pricing_grille(wb)

    wb.save(DST)
    print(f"✓ Generated {DST}")
    print(f"  Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
