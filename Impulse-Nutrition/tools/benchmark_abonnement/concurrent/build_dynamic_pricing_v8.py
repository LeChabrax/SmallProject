"""
Build v8 — 2 pricing tabs :
  1. "Pricing Grille C"     → 3 tables côte à côte (variation C des previews)
  2. "Pricing Dashboard A"  → dashboard top + tables stackées (variation A)

Both share the same cascade logic :
  Concentrate dropdown → Concentrate retenu
      → Isolate PP OPT (via anti-cannib)
          → Isolate options → Isolate retenu
              → Recovery PP OPT → Recovery options → Recovery retenu

Same hypothèses source (PR, TVA, remise) at the top of each tab.

Design : Bloomberg-meets-Notion. Charbon + ambre, couleurs par produit (bleu
Concentrate / vert Isolate / rouge Recovery). Mono pour les nombres, typo
sérieuse pour les labels, amber fort pour les sélections.
"""
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = Path("/Users/antoinechabrat/Downloads/Copie de benchmark_whey_concurrent 5 1.xlsx")
DST = Path(
    "/Users/antoinechabrat/Documents/SmallProject/Impulse-Nutrition/benchmark/"
    "concurrent/benchmark_whey_concurrent_v8_dynamic.xlsx"
)

# ─── Design tokens ──────────────────────────────────────────────────────
INK       = "0F172A"
MUTED     = "64748B"
BORDER    = "CBD5E1"
LIGHT     = "F1F5F9"
WHITE     = "FFFFFF"
ACCENT    = "DC2626"
AMBER     = "FEF3C7"
AMBER_STR = "F59E0B"

COL_CONC  = "1D4ED8"  # blue-700
COL_ISO   = "047857"  # emerald-700
COL_REC   = "B91C1C"  # red-700

FONT_TITLE     = Font(name="Arial", size=20, bold=True, color=INK)
FONT_SUBTITLE  = Font(name="Arial", size=10, italic=True, color=MUTED)
FONT_SECTION   = Font(name="Arial", size=13, bold=True, color=WHITE)
FONT_LABEL     = Font(name="Arial", size=9, bold=True, color=INK)
FONT_LABEL_MUT = Font(name="Arial", size=9, italic=True, color=MUTED)
FONT_HEADER    = Font(name="Arial", size=9, bold=True, color=WHITE)
FONT_NUM       = Font(name="Consolas", size=10, color=INK)
FONT_NUM_BOLD  = Font(name="Consolas", size=11, bold=True, color=INK)
FONT_PRIX_BIG  = Font(name="Arial", size=20, bold=True, color=ACCENT)
FONT_MARGE_BIG = Font(name="Arial", size=14, bold=True, color=INK)
FONT_FLAG      = Font(name="Arial", size=11, color=INK)
FONT_DROPDOWN  = Font(name="Arial", size=14, bold=True, color=INK)

FILL_CONC      = PatternFill(start_color=COL_CONC, end_color=COL_CONC, fill_type="solid")
FILL_ISO       = PatternFill(start_color=COL_ISO, end_color=COL_ISO, fill_type="solid")
FILL_REC       = PatternFill(start_color=COL_REC, end_color=COL_REC, fill_type="solid")
FILL_MUTED     = PatternFill(start_color=MUTED, end_color=MUTED, fill_type="solid")
FILL_LIGHT     = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
FILL_AMBER     = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
FILL_AMBER_STR = PatternFill(start_color=AMBER_STR, end_color=AMBER_STR, fill_type="solid")
FILL_ACCENT    = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
FILL_WHITE     = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

THIN_BORDER  = Side(style="thin", color=BORDER)
BOX_BORDER   = Border(left=THIN_BORDER, right=THIN_BORDER, top=THIN_BORDER, bottom=THIN_BORDER)

NUM_EUR = '#,##0.00" €"'
NUM_PCT = "0.0%"
NUM_PCT_SIGNED = '+0.0%;-0.0%;0.0%'

# Benchmark refs (same as v7)
CONC_BENCH = [
    "'Synthèse prix'!H14", "'Synthèse prix'!H15", "'Synthèse prix'!H19",
    "'Synthèse prix'!H20", "'Synthèse prix'!H24", "'Synthèse prix'!H25",
    "'Synthèse prix'!H26", "'Synthèse prix'!H31", "'Synthèse prix'!H37",
    "'Synthèse prix'!H50",
]
ISO_BENCH = [
    "'Synthèse prix'!H7", "'Synthèse prix'!H10", "'Synthèse prix'!H11",
    "'Synthèse prix'!H18", "'Synthèse prix'!H21", "'Synthèse prix'!H22",
    "'Synthèse prix'!H23", "'Synthèse prix'!H29", "'Synthèse prix'!H30",
    "'Synthèse prix'!H35", "'Synthèse prix'!H41", "'Synthèse prix'!H47",
]
REC_BENCH = [
    "'Synthèse prix'!H5", "'Synthèse prix'!H6", "'Synthèse prix'!H32",
]

OFFSETS = [-4, -3, -2, -1, 0, 1, 2, 3, 5, 7]
CONC_PRICES = [29.9, 31.9, 33.9, 34.9, 36.9, 37.9, 39.9, 41.9, 42.9, 44.9]


def add_dropdown(ws, cell_ref):
    dv = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(cell_ref)
    ws[cell_ref] = 5
    ws[cell_ref].font = FONT_DROPDOWN
    ws[cell_ref].fill = FILL_AMBER_STR
    ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center")
    ws[cell_ref].border = BOX_BORDER


def flag_formula_concentrate(marge_cell):
    return (
        f'=IF({marge_cell}<0.30,"⚠️ Marge basse",'
        f'IF({marge_cell}<0.35,"🟡 Limite",'
        f'IF({marge_cell}<=0.42,"🎯 Optimal","🔵 Premium")))'
    )


def flag_formula_cascading(marge_cell, prix_cell, cannib_ref, threshold=0.10):
    return (
        f'=IF({marge_cell}<0.30,"⚠️ Marge basse",'
        f'IF(IFERROR(({prix_cell}-{cannib_ref})/{cannib_ref},0)<{threshold},"🚫 Cannibalisme",'
        f'IF({marge_cell}<0.35,"🟡 Limite",'
        f'IF({marge_cell}<=0.42,"🎯 Optimal","🔵 Premium"))))'
    )


# ═════════════════════════════════════════════════════════════════════════
#  VARIATION C — GRILLE 3 COLONNES CÔTE À CÔTE
# ═════════════════════════════════════════════════════════════════════════
def build_variation_c(wb):
    """3 tables side-by-side. 14 columns wide.

    Column layout :
       A-D  : Concentrate (# | Prix | Marge | Flag)
       E    : gap
       F-I  : Isolate
       J    : gap
       K-N  : Recovery
    """
    if "Pricing Grille C" in wb.sheetnames:
        del wb["Pricing Grille C"]
    ws = wb.create_sheet("Pricing Grille C", 0)

    MAX_COL = 14

    # ── Row 1 : Title
    ws.merge_cells("A1:N1")
    ws["A1"] = "PRICING DYNAMIQUE · GRILLE 3 COLONNES"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    # ── Row 2 : Subtitle
    ws.merge_cells("A2:N2")
    ws["A2"] = "Les 3 produits en parallèle · 1 sélection Concentrate → Isolate & Recovery se recalculent"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Row 4 : Hypothèses label
    ws["A4"] = "HYPOTHÈSES PARTAGÉES"
    ws["A4"].font = Font(name="Arial", size=9, bold=True, color=MUTED)
    ws.merge_cells("A4:N4")

    # ── Row 5 : Inline hypothèses
    hyp = [
        (1, "TVA", 0.055, NUM_PCT),
        (3, "Remise", 0.18, NUM_PCT),
        (5, "PR Conc", 21, NUM_EUR),
        (7, "PR Iso", 25, NUM_EUR),
        (9, "PR Rec", 12.15, NUM_EUR),
    ]
    for col, label, value, fmt in hyp:
        c = ws.cell(row=5, column=col, value=label)
        c.font = FONT_LABEL_MUT
        c.alignment = Alignment(horizontal="right")
        v = ws.cell(row=5, column=col + 1, value=value)
        v.number_format = fmt
        v.font = FONT_NUM_BOLD
        v.fill = FILL_LIGHT
        v.alignment = Alignment(horizontal="center")
    # Named hypothèse cells
    tva_cell    = "$B$5"
    remise_cell = "$D$5"
    pr_conc     = "$F$5"
    pr_iso      = "$H$5"
    pr_rec      = "$J$5"

    # ── Row 7 : Section headers (merged per product)
    ws.merge_cells("A7:D7")
    ws["A7"] = "▌ CONCENTRATE"
    ws["A7"].font = FONT_SECTION
    ws["A7"].fill = FILL_CONC
    ws["A7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[7].height = 28

    ws.merge_cells("F7:I7")
    ws["F7"] = "▌ ISOLATE   ⬅ cascade"
    ws["F7"].font = FONT_SECTION
    ws["F7"].fill = FILL_ISO
    ws["F7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.merge_cells("K7:N7")
    ws["K7"] = "▌ RECOVERY   ⬅ cascade"
    ws["K7"].font = FONT_SECTION
    ws["K7"].fill = FILL_REC
    ws["K7"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Row 9 : "🎯 Option" label
    for col in [1, 6, 11]:
        c = ws.cell(row=9, column=col, value="🎯 Option sélectionnée")
        c.font = FONT_LABEL
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 3)

    # ── Row 10 : DROPDOWNS
    conc_dropdown = "B10"
    iso_dropdown  = "G10"
    rec_dropdown  = "L10"
    for cell_ref in [conc_dropdown, iso_dropdown, rec_dropdown]:
        add_dropdown(ws, cell_ref)
    ws.row_dimensions[10].height = 30
    # Merge dropdown cells across 3 cols each
    ws.merge_cells("B10:D10")
    ws.merge_cells("G10:I10")
    ws.merge_cells("L10:N10")

    # ── Row 12 : "Prix retenu" label
    for col in [1, 6, 11]:
        c = ws.cell(row=12, column=col, value="→ Prix retenu")
        c.font = FONT_LABEL_MUT
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=12, start_column=col, end_row=12, end_column=col + 3)

    # ── Row 13 : BIG prix retenu
    options_start_row = 29
    options_end_row = 38

    conc_retenu_cell = "B13"
    iso_retenu_cell  = "G13"
    rec_retenu_cell  = "L13"

    ws[conc_retenu_cell] = f"=INDEX($B${options_start_row}:$B${options_end_row},${conc_dropdown})"
    ws[iso_retenu_cell]  = f"=INDEX($G${options_start_row}:$G${options_end_row},${iso_dropdown})"
    ws[rec_retenu_cell]  = f"=INDEX($L${options_start_row}:$L${options_end_row},${rec_dropdown})"

    for cell_ref, fill in [(conc_retenu_cell, FILL_AMBER), (iso_retenu_cell, FILL_AMBER), (rec_retenu_cell, FILL_AMBER)]:
        c = ws[cell_ref]
        c.font = FONT_PRIX_BIG
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = NUM_EUR
        c.border = BOX_BORDER
    # Merge big prix across cols
    ws.merge_cells("B13:D13")
    ws.merge_cells("G13:I13")
    ws.merge_cells("L13:N13")
    ws.row_dimensions[13].height = 36

    # ── Row 14 : Marge retenue
    for col_letter, pr_cell, retenu_cell, merge_start, merge_end in [
        ("B", pr_conc, conc_retenu_cell, "A14", "D14"),
        ("G", pr_iso, iso_retenu_cell, "F14", "I14"),
        ("L", pr_rec, rec_retenu_cell, "K14", "N14"),
    ]:
        # Label
        label_col = merge_start[0]
        label_cell = f"{label_col}14"
        ws[label_cell] = "Marge nette"
        ws[label_cell].font = FONT_LABEL_MUT
        ws[label_cell].alignment = Alignment(horizontal="left", indent=1)
        # Value in the col + 1
        val_col = chr(ord(label_col) + 1)
        val_cell = f"{val_col}14"
        ws[val_cell] = (
            f"=IFERROR(({retenu_cell}*(1-{remise_cell})-{pr_cell})/({retenu_cell}*(1-{remise_cell})),0)"
        )
        ws[val_cell].font = FONT_MARGE_BIG
        ws[val_cell].number_format = NUM_PCT
        ws[val_cell].alignment = Alignment(horizontal="center", vertical="center")
        # Merge across 3 cols for the value cell
        ws.merge_cells(start_row=14, start_column=ord(val_col) - 64, end_row=14, end_column=ord(merge_end[0]) - 64)
    ws.row_dimensions[14].height = 22

    # ── Row 15 : (empty)

    # ── Row 16 : BENCHMARK label
    for col in [1, 6, 11]:
        c = ws.cell(row=16, column=col, value="BENCHMARK MARCHÉ")
        c.font = Font(name="Arial", size=8, bold=True, color=MUTED)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=16, start_column=col, end_row=16, end_column=col + 3)

    # ── Row 17 : Moy | Méd | Min | Max (4 values across 4 cols)
    for col, bench in [(1, CONC_BENCH), (6, ISO_BENCH), (11, REC_BENCH)]:
        bench_list = ",".join(bench)
        stats = [("Moy", f"=AVERAGE({bench_list})"), ("Méd", f"=MEDIAN({bench_list})"),
                 ("Min", f"=MIN({bench_list})"), ("Max", f"=MAX({bench_list})")]
        for i, (label, formula) in enumerate(stats):
            c = ws.cell(row=17, column=col + i)
            c.value = formula
            c.number_format = NUM_EUR
            c.font = FONT_NUM
            c.alignment = Alignment(horizontal="center")
        # Sub-labels below
        for i, (label, _) in enumerate(stats):
            lbl = ws.cell(row=18, column=col + i, value=label)
            lbl.font = Font(name="Arial", size=8, italic=True, color=MUTED)
            lbl.alignment = Alignment(horizontal="center")
    # Store median refs for use in options
    conc_median_cell = "B17"  # first stat col of Concentrate = Moy, so Méd is C17
    conc_median_cell = "C17"
    iso_median_cell = "G17"
    rec_median_cell = "L17"

    # ── Row 20 : CALCUL AUTO header (Isolate/Recovery only)
    for col in [6, 11]:
        c = ws.cell(row=20, column=col, value="CALCUL AUTO — Algo Option C")
        c.font = Font(name="Arial", size=8, bold=True, color=MUTED)
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=20, start_column=col, end_row=20, end_column=col + 3)

    # ── Rows 21-25 : Calcul auto values (5 bornes + PP OPT)
    # Isolate
    iso_bench_list = ",".join(ISO_BENCH)
    iso_calc = [
        ("Plancher marge 38%",  f"=CEILING({pr_iso}/(1-0.38)/(1-{remise_cell})+0.1,1)-0.1"),
        ("Cible marge 40%",     f"=CEILING({pr_iso}/(1-0.40)/(1-{remise_cell})+0.1,1)-0.1"),
        ("Anti-cannib (C×1,10)", f"=CEILING({conc_retenu_cell}*1.10+0.1,1)-0.1"),
        ("Marché médiane",      f"=MEDIAN({iso_bench_list})"),
    ]
    for i, (label, formula) in enumerate(iso_calc):
        r = 21 + i
        lbl_cell = ws.cell(row=r, column=6, value=label)
        lbl_cell.font = FONT_LABEL_MUT
        lbl_cell.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        val_cell = ws.cell(row=r, column=9, value=formula)
        val_cell.font = FONT_NUM
        val_cell.number_format = NUM_EUR
        val_cell.alignment = Alignment(horizontal="center")
    # PP OPT row 25
    iso_pp_opt_cell = "I25"
    ws["F25"] = "→ PP OPTIMAL"
    ws["F25"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws["F25"].fill = FILL_ACCENT
    ws["F25"].alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("F25:H25")
    ws[iso_pp_opt_cell] = f"=MAX(I21,I23,MIN(I22,I24))"
    ws[iso_pp_opt_cell].font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws[iso_pp_opt_cell].fill = FILL_ACCENT
    ws[iso_pp_opt_cell].number_format = NUM_EUR
    ws[iso_pp_opt_cell].alignment = Alignment(horizontal="center", vertical="center")

    # Recovery
    rec_calc = [
        ("Plancher marge 38%",  f"=CEILING({pr_rec}/(1-0.38)/(1-{remise_cell})+0.1,1)-0.1"),
        ("Plancher C+5%",       f"=CEILING({conc_retenu_cell}*1.05+0.1,1)-0.1"),
        ("Plafond I−5%",        f"=CEILING({iso_retenu_cell}*0.95+0.1,1)-0.1"),
        ("Marché réf (Impulse)", "='Synthèse prix'!H5"),
    ]
    for i, (label, formula) in enumerate(rec_calc):
        r = 21 + i
        lbl_cell = ws.cell(row=r, column=11, value=label)
        lbl_cell.font = FONT_LABEL_MUT
        lbl_cell.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=13)
        val_cell = ws.cell(row=r, column=14, value=formula)
        val_cell.font = FONT_NUM
        val_cell.number_format = NUM_EUR
        val_cell.alignment = Alignment(horizontal="center")
    # PP OPT row 25 (Recovery)
    rec_pp_opt_cell = "N25"
    ws["K25"] = "→ PP OPTIMAL"
    ws["K25"].font = Font(name="Arial", size=10, bold=True, color=WHITE)
    ws["K25"].fill = FILL_ACCENT
    ws["K25"].alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells("K25:M25")
    ws[rec_pp_opt_cell] = f"=MAX(N21,N22,MIN(N23,N24))"
    ws[rec_pp_opt_cell].font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws[rec_pp_opt_cell].fill = FILL_ACCENT
    ws[rec_pp_opt_cell].number_format = NUM_EUR
    ws[rec_pp_opt_cell].alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 27 : Table headers (# | Prix | Marge | Flag) × 3
    headers = ["#", "Prix", "Marge %", "Flag"]
    for group_start, fill in [(1, FILL_MUTED), (6, FILL_MUTED), (11, FILL_MUTED)]:
        for i, h in enumerate(headers):
            c = ws.cell(row=27, column=group_start + i, value=h)
            c.font = FONT_HEADER
            c.fill = fill
            c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[27].height = 22

    # ── Rows 29-38 : Options (Concentrate hardcoded, Isolate/Recovery dynamic)
    for i in range(10):
        r = options_start_row + i
        opt_num = i + 1

        # Concentrate (cols A-D)
        ws.cell(row=r, column=1, value=opt_num).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=CONC_PRICES[i]).number_format = NUM_EUR
        ws.cell(row=r, column=3, value=(
            f"=IFERROR((B{r}*(1-{remise_cell})-{pr_conc})/(B{r}*(1-{remise_cell})),0)"
        )).number_format = NUM_PCT
        ws.cell(row=r, column=4, value=flag_formula_concentrate(f"C{r}"))
        ws.cell(row=r, column=4).alignment = Alignment(horizontal="left", indent=1)

        # Isolate (cols F-I) — dynamic prices from I25 (PP OPT)
        offset = OFFSETS[i]
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        prix_expr = f"{iso_pp_opt_cell}" if offset == 0 else f"{iso_pp_opt_cell}{sign}{abs_off}"
        ws.cell(row=r, column=6, value=opt_num).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=7, value=f"=CEILING({prix_expr}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=r, column=8, value=(
            f"=IFERROR((G{r}*(1-{remise_cell})-{pr_iso})/(G{r}*(1-{remise_cell})),0)"
        )).number_format = NUM_PCT
        ws.cell(row=r, column=9, value=flag_formula_cascading(f"H{r}", f"G{r}", conc_retenu_cell, 0.10))
        ws.cell(row=r, column=9).alignment = Alignment(horizontal="left", indent=1)

        # Recovery (cols K-N) — dynamic prices from N25
        prix_expr_rec = f"{rec_pp_opt_cell}" if offset == 0 else f"{rec_pp_opt_cell}{sign}{abs_off}"
        ws.cell(row=r, column=11, value=opt_num).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=r, column=11).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=12, value=f"=CEILING({prix_expr_rec}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=r, column=13, value=(
            f"=IFERROR((L{r}*(1-{remise_cell})-{pr_rec})/(L{r}*(1-{remise_cell})),0)"
        )).number_format = NUM_PCT
        ws.cell(row=r, column=14, value=flag_formula_cascading(f"M{r}", f"L{r}", conc_retenu_cell, 0.05))
        ws.cell(row=r, column=14).alignment = Alignment(horizontal="left", indent=1)

        # Font + alignment numeric cells
        for col in [2, 3, 7, 8, 12, 13]:
            ws.cell(row=r, column=col).font = FONT_NUM
            ws.cell(row=r, column=col).alignment = Alignment(horizontal="right", indent=1)

        # Zebra rows (alternating) — skip gap cols E, J
        if i % 2 == 1:
            for col in [1, 2, 3, 4, 6, 7, 8, 9, 11, 12, 13, 14]:
                c = ws.cell(row=r, column=col)
                if not c.fill.start_color or c.fill.start_color.value in (None, "00000000"):
                    c.fill = FILL_LIGHT

        # Highlight the center option (#5) across all 3 products
        if opt_num == 5:
            for col in [1, 2, 3, 4, 6, 7, 8, 9, 11, 12, 13, 14]:
                c = ws.cell(row=r, column=col)
                c.fill = FILL_AMBER

    # Column widths
    widths = {
        "A": 5, "B": 13, "C": 12, "D": 20,
        "E": 2,
        "F": 5, "G": 13, "H": 12, "I": 20,
        "J": 2,
        "K": 5, "L": 13, "M": 12, "N": 20,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A7"


# ═════════════════════════════════════════════════════════════════════════
#  VARIATION A — DASHBOARD TOP + TABLES STACKED
# ═════════════════════════════════════════════════════════════════════════
def build_variation_a(wb):
    """Dashboard header + stacked tables (like v7 but with summary row on top)."""
    if "Pricing Dashboard A" in wb.sheetnames:
        del wb["Pricing Dashboard A"]
    ws = wb.create_sheet("Pricing Dashboard A", 1)

    # ── Row 1-2 : Title + subtitle
    ws.merge_cells("A1:I1")
    ws["A1"] = "PRICING DYNAMIQUE · DASHBOARD GAMME"
    ws["A1"].font = FONT_TITLE
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:I2")
    ws["A2"] = "Vue gamme en header · tables stackées en dessous"
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Rows 4-9 : Hypothèses (vertical)
    ws["A4"] = "HYPOTHÈSES PARTAGÉES"
    ws["A4"].font = Font(name="Arial", size=9, bold=True, color=MUTED)
    ws.merge_cells("A4:I4")
    hyp_rows = [
        (5, "TVA (compléments alim.)", 0.055, NUM_PCT),
        (6, "Remise client moyenne", 0.18, NUM_PCT),
        (7, "PR Concentrate (€ HT)", 21, NUM_EUR),
        (8, "PR Isolate (€ HT)", 25, NUM_EUR),
        (9, "PR Recovery (€ HT)", 12.15, NUM_EUR),
    ]
    for row, label, value, fmt in hyp_rows:
        ws.cell(row=row, column=1, value=label).font = FONT_LABEL
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", indent=1)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = FONT_NUM_BOLD
        c.fill = FILL_LIGHT
        c.alignment = Alignment(horizontal="center")

    tva_cell    = "$B$5"
    remise_cell = "$B$6"
    pr_conc     = "$B$7"
    pr_iso      = "$B$8"
    pr_rec      = "$B$9"

    # ── Rows 11-16 : DASHBOARD
    ws.merge_cells("A11:I11")
    ws["A11"] = "▌ DASHBOARD GAMME"
    ws["A11"].font = FONT_SECTION
    ws["A11"].fill = PatternFill(start_color=INK, end_color=INK, fill_type="solid")
    ws["A11"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[11].height = 28

    # Row 12 : labels
    ws["A12"] = ""
    ws["B12"] = "CONCENTRATE"
    ws["B12"].font = Font(name="Arial", size=11, bold=True, color=COL_CONC)
    ws["B12"].alignment = Alignment(horizontal="center")
    ws.merge_cells("B12:C12")

    ws["D12"] = "ISOLATE"
    ws["D12"].font = Font(name="Arial", size=11, bold=True, color=COL_ISO)
    ws["D12"].alignment = Alignment(horizontal="center")
    ws.merge_cells("D12:E12")

    ws["F12"] = "RECOVERY"
    ws["F12"].font = Font(name="Arial", size=11, bold=True, color=COL_REC)
    ws["F12"].alignment = Alignment(horizontal="center")
    ws.merge_cells("F12:G12")

    ws["H12"] = "ÉCART GAMME"
    ws["H12"].font = Font(name="Arial", size=11, bold=True, color=MUTED)
    ws["H12"].alignment = Alignment(horizontal="center")
    ws.merge_cells("H12:I12")

    # Forward references to tables below (rows of the sélection row of each table)
    CONC_RETENU = "$D$34"   # Concentrate sélection row
    ISO_RETENU  = "$D$55"   # Isolate sélection row
    REC_RETENU  = "$D$76"   # Recovery sélection row

    # Row 13 : Prix retenus
    ws["A13"] = "Prix retenu"
    ws["A13"].font = FONT_LABEL
    ws["A13"].alignment = Alignment(horizontal="left", indent=1)

    for label_col, retenu_ref, merge_end in [("B", CONC_RETENU, "C"), ("D", ISO_RETENU, "E"), ("F", REC_RETENU, "G")]:
        c = ws[f"{label_col}13"]
        c.value = f"={retenu_ref}"
        c.font = FONT_PRIX_BIG
        c.fill = FILL_AMBER
        c.number_format = NUM_EUR
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX_BORDER
        ws.merge_cells(f"{label_col}13:{merge_end}13")
    # Écart total gamme (Isolate vs Concentrate)
    ws["H13"] = f"=IFERROR(({ISO_RETENU}-{CONC_RETENU})/{CONC_RETENU},0)"
    ws["H13"].number_format = NUM_PCT_SIGNED
    ws["H13"].font = FONT_MARGE_BIG
    ws["H13"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("H13:I13")
    ws.row_dimensions[13].height = 36

    # Row 14 : Marges
    ws["A14"] = "Marge nette"
    ws["A14"].font = FONT_LABEL
    ws["A14"].alignment = Alignment(horizontal="left", indent=1)
    for label_col, retenu_ref, pr_ref, merge_end in [
        ("B", CONC_RETENU, pr_conc, "C"),
        ("D", ISO_RETENU, pr_iso, "E"),
        ("F", REC_RETENU, pr_rec, "G"),
    ]:
        c = ws[f"{label_col}14"]
        c.value = (
            f"=IFERROR(({retenu_ref}*(1-{remise_cell})-{pr_ref})/({retenu_ref}*(1-{remise_cell})),0)"
        )
        c.number_format = NUM_PCT
        c.font = FONT_MARGE_BIG
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{label_col}14:{merge_end}14")

    ws["H14"] = f"=IFERROR(({REC_RETENU}-{CONC_RETENU})/{CONC_RETENU},0)"
    ws["H14"].number_format = NUM_PCT_SIGNED
    ws["H14"].font = Font(name="Arial", size=11, color=MUTED)
    ws["H14"].alignment = Alignment(horizontal="center")
    ws.merge_cells("H14:I14")

    # Row 15 : Flag (use same flag logic inline)
    ws["A15"] = "Recommandation"
    ws["A15"].font = FONT_LABEL
    ws["A15"].alignment = Alignment(horizontal="left", indent=1)
    for label_col, retenu_ref, pr_ref, merge_end in [
        ("B", CONC_RETENU, pr_conc, "C"),
        ("D", ISO_RETENU, pr_iso, "E"),
        ("F", REC_RETENU, pr_rec, "G"),
    ]:
        marge_expr = (
            f"IFERROR(({retenu_ref}*(1-{remise_cell})-{pr_ref})/({retenu_ref}*(1-{remise_cell})),0)"
        )
        c = ws[f"{label_col}15"]
        c.value = (
            f'=IF({marge_expr}<0.30,"⚠️ Marge basse",'
            f'IF({marge_expr}<0.35,"🟡 Limite",'
            f'IF({marge_expr}<=0.42,"🎯 Optimal","🔵 Premium")))'
        )
        c.font = FONT_FLAG
        c.alignment = Alignment(horizontal="center")
        ws.merge_cells(f"{label_col}15:{merge_end}15")

    # Row 16 : empty
    ws.row_dimensions[16].height = 12

    # ── Stacked tables
    # Table 1 : Concentrate (rows 18-34, selection at 34)
    build_stacked_table_conc(ws, 18, tva_cell, remise_cell, pr_conc)
    # Table 2 : Isolate (rows 36-55, selection at 55)
    build_stacked_table_iso(ws, 36, tva_cell, remise_cell, pr_iso, CONC_RETENU)
    # Table 3 : Recovery (rows 57-76, selection at 76)
    build_stacked_table_rec(ws, 57, tva_cell, remise_cell, pr_rec, CONC_RETENU, ISO_RETENU)

    # Column widths
    widths = {"A": 22, "B": 12, "C": 12, "D": 14, "E": 11, "F": 13, "G": 16, "H": 16, "I": 22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    ws.freeze_panes = "A17"


def build_stacked_table_conc(ws, start, tva_cell, remise_cell, pr_cell):
    r = start
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value="▌ TABLEAU 1 — CONCENTRATE")
    c.font = FONT_SECTION
    c.fill = FILL_CONC
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26

    # Benchmark stats row
    r += 2
    bench_list = ",".join(CONC_BENCH)
    ws.cell(row=r, column=1, value="Benchmark").font = FONT_LABEL_MUT
    stats = [("Moy", f"=AVERAGE({bench_list})"), ("Méd", f"=MEDIAN({bench_list})"),
             ("Min", f"=MIN({bench_list})"), ("Max", f"=MAX({bench_list})")]
    for i, (label, formula) in enumerate(stats):
        lbl = ws.cell(row=r, column=2 + i * 2, value=label)
        lbl.font = Font(name="Arial", size=8, italic=True, color=MUTED)
        lbl.alignment = Alignment(horizontal="right")
        val = ws.cell(row=r, column=3 + i * 2, value=formula)
        val.font = FONT_NUM
        val.number_format = NUM_EUR
        val.alignment = Alignment(horizontal="left", indent=1)
    median_cell = f"E{r}"  # Méd value col

    r += 2
    # Headers
    headers = ["#", "PP TTC", "PP HT", "PVM net", "Marge €", "% Marge", "vs Médiane", "—", "Flag"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_MUTED
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 20

    r += 1
    # 10 options
    for i in range(10):
        row = r + i
        ws.cell(row=row, column=1, value=i + 1).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=CONC_PRICES[i]).number_format = NUM_EUR
        ws.cell(row=row, column=3, value=f"=B{row}/(1+{tva_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=4, value=f"=B{row}*(1-{remise_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=5, value=f"=D{row}-{pr_cell}").number_format = NUM_EUR
        ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/D{row},0)").number_format = NUM_PCT
        ws.cell(row=row, column=7, value=f"=IFERROR((B{row}-{median_cell})/{median_cell},0)").number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=8, value="—").alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=9, value=flag_formula_concentrate(f"F{row}"))
        for col in range(2, 8):
            ws.cell(row=row, column=col).font = FONT_NUM
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", indent=1)
        if i % 2 == 1:
            for col in range(1, 10):
                c = ws.cell(row=row, column=col)
                if not c.fill.start_color or c.fill.start_color.value in (None, "00000000"):
                    c.fill = FILL_LIGHT
    options_start = r
    options_end = r + 9

    # Sélection row (below options)
    r += 11
    ws.cell(row=r, column=1, value="🎯 Option sélectionnée").font = FONT_LABEL
    sel_cell = f"B{r}"
    add_dropdown(ws, sel_cell)
    ws.row_dimensions[r].height = 26
    ws.cell(row=r, column=3, value="→ Prix retenu").font = FONT_LABEL
    ws.cell(row=r, column=4, value=f"=INDEX($B${options_start}:$B${options_end},${sel_cell})")
    ws.cell(row=r, column=4).font = FONT_PRIX_BIG
    ws.cell(row=r, column=4).fill = FILL_AMBER
    ws.cell(row=r, column=4).number_format = NUM_EUR
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5, value="Marge").font = FONT_LABEL
    ws.cell(row=r, column=6, value=f"=INDEX($F${options_start}:$F${options_end},${sel_cell})")
    ws.cell(row=r, column=6).font = FONT_MARGE_BIG
    ws.cell(row=r, column=6).fill = FILL_AMBER
    ws.cell(row=r, column=6).number_format = NUM_PCT
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")


def build_stacked_table_iso(ws, start, tva_cell, remise_cell, pr_cell, concentrate_retenu_cell):
    r = start
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value="▌ TABLEAU 2 — ISOLATE (cascade auto)")
    c.font = FONT_SECTION
    c.fill = FILL_ISO
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26

    # Calcul auto row (compact 1 line)
    r += 2
    labels = [
        (1, "Plancher 38%"),
        (3, "Cible 40%"),
        (5, "Anti-cannib"),
        (7, "Marché méd."),
    ]
    for col, lbl in labels:
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = Font(name="Arial", size=8, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=9, value="→ PP OPT").font = Font(name="Arial", size=9, bold=True, color=ACCENT)
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")

    r += 1
    iso_bench_list = ",".join(ISO_BENCH)
    ws.cell(row=r, column=2, value=f"=CEILING({pr_cell}/(1-0.38)/(1-{remise_cell})+0.1,1)-0.1").number_format = NUM_EUR
    marge38_cell = f"B{r}"
    ws.cell(row=r, column=4, value=f"=CEILING({pr_cell}/(1-0.40)/(1-{remise_cell})+0.1,1)-0.1").number_format = NUM_EUR
    marge40_cell = f"D{r}"
    ws.cell(row=r, column=6, value=f"=CEILING({concentrate_retenu_cell}*1.10+0.1,1)-0.1").number_format = NUM_EUR
    anticannib_cell = f"F{r}"
    ws.cell(row=r, column=8, value=f"=MEDIAN({iso_bench_list})").number_format = NUM_EUR
    marche_cell = f"H{r}"
    pp_opt_cell = f"I{r}"
    ws.cell(row=r, column=9, value=f"=MAX({marge38_cell},{anticannib_cell},MIN({marge40_cell},{marche_cell}))")
    ws.cell(row=r, column=9).font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws.cell(row=r, column=9).fill = FILL_ACCENT
    ws.cell(row=r, column=9).number_format = NUM_EUR
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
    for col in [2, 4, 6, 8]:
        ws.cell(row=r, column=col).font = FONT_NUM
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    r += 2
    # Benchmark stats row
    ws.cell(row=r, column=1, value="Benchmark").font = FONT_LABEL_MUT
    stats = [("Moy", f"=AVERAGE({iso_bench_list})"), ("Méd", f"=MEDIAN({iso_bench_list})"),
             ("Min", f"=MIN({iso_bench_list})"), ("Max", f"=MAX({iso_bench_list})")]
    for i, (label, formula) in enumerate(stats):
        lbl = ws.cell(row=r, column=2 + i * 2, value=label)
        lbl.font = Font(name="Arial", size=8, italic=True, color=MUTED)
        lbl.alignment = Alignment(horizontal="right")
        val = ws.cell(row=r, column=3 + i * 2, value=formula)
        val.font = FONT_NUM
        val.number_format = NUM_EUR
        val.alignment = Alignment(horizontal="left", indent=1)
    median_cell = f"E{r}"

    r += 2
    # Headers
    headers = ["#", "PP TTC", "PP HT", "PVM net", "Marge €", "% Marge", "vs Médiane", "vs Concentrate", "Flag"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_MUTED
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 20

    r += 1
    options_start = r
    for i, offset in enumerate(OFFSETS):
        row = r + i
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"{pp_opt_cell}" if offset == 0 else f"{pp_opt_cell}{sign}{abs_off}"
        ws.cell(row=row, column=1, value=i + 1).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=f"=CEILING({expr}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=row, column=3, value=f"=B{row}/(1+{tva_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=4, value=f"=B{row}*(1-{remise_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=5, value=f"=D{row}-{pr_cell}").number_format = NUM_EUR
        ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/D{row},0)").number_format = NUM_PCT
        ws.cell(row=row, column=7, value=f"=IFERROR((B{row}-{median_cell})/{median_cell},0)").number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=8, value=f"=IFERROR((B{row}-{concentrate_retenu_cell})/{concentrate_retenu_cell},0)").number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=9, value=flag_formula_cascading(f"F{row}", f"B{row}", concentrate_retenu_cell, 0.10))
        for col in range(2, 9):
            ws.cell(row=row, column=col).font = FONT_NUM
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", indent=1)
        if i % 2 == 1:
            for col in range(1, 10):
                c = ws.cell(row=row, column=col)
                if not c.fill.start_color or c.fill.start_color.value in (None, "00000000"):
                    c.fill = FILL_LIGHT
    options_end = r + 9

    r += 11
    ws.cell(row=r, column=1, value="🎯 Option sélectionnée").font = FONT_LABEL
    sel_cell = f"B{r}"
    add_dropdown(ws, sel_cell)
    ws.row_dimensions[r].height = 26
    ws.cell(row=r, column=3, value="→ Prix retenu").font = FONT_LABEL
    ws.cell(row=r, column=4, value=f"=INDEX($B${options_start}:$B${options_end},${sel_cell})")
    ws.cell(row=r, column=4).font = FONT_PRIX_BIG
    ws.cell(row=r, column=4).fill = FILL_AMBER
    ws.cell(row=r, column=4).number_format = NUM_EUR
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5, value="Marge").font = FONT_LABEL
    ws.cell(row=r, column=6, value=f"=INDEX($F${options_start}:$F${options_end},${sel_cell})")
    ws.cell(row=r, column=6).font = FONT_MARGE_BIG
    ws.cell(row=r, column=6).fill = FILL_AMBER
    ws.cell(row=r, column=6).number_format = NUM_PCT
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")


def build_stacked_table_rec(ws, start, tva_cell, remise_cell, pr_cell, concentrate_retenu_cell, isolate_retenu_cell):
    r = start
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    c = ws.cell(row=r, column=1, value="▌ TABLEAU 3 — RECOVERY (cascade auto)")
    c.font = FONT_SECTION
    c.fill = FILL_REC
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[r].height = 26

    r += 2
    labels = [
        (1, "Plancher 38%"),
        (3, "Plancher C+5%"),
        (5, "Plafond I−5%"),
        (7, "Marché réf"),
    ]
    for col, lbl in labels:
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = Font(name="Arial", size=8, italic=True, color=MUTED)
        c.alignment = Alignment(horizontal="right", indent=1)
    ws.cell(row=r, column=9, value="→ PP OPT").font = Font(name="Arial", size=9, bold=True, color=ACCENT)
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center")

    r += 1
    rec_bench_list = ",".join(REC_BENCH)
    ws.cell(row=r, column=2, value=f"=CEILING({pr_cell}/(1-0.38)/(1-{remise_cell})+0.1,1)-0.1").number_format = NUM_EUR
    marge38_cell = f"B{r}"
    ws.cell(row=r, column=4, value=f"=CEILING({concentrate_retenu_cell}*1.05+0.1,1)-0.1").number_format = NUM_EUR
    floor_cell = f"D{r}"
    ws.cell(row=r, column=6, value=f"=CEILING({isolate_retenu_cell}*0.95+0.1,1)-0.1").number_format = NUM_EUR
    ceiling_cell = f"F{r}"
    ws.cell(row=r, column=8, value="='Synthèse prix'!H5").number_format = NUM_EUR
    marche_cell = f"H{r}"
    pp_opt_cell = f"I{r}"
    ws.cell(row=r, column=9, value=f"=MAX({marge38_cell},{floor_cell},MIN({ceiling_cell},{marche_cell}))")
    ws.cell(row=r, column=9).font = Font(name="Arial", size=12, bold=True, color=WHITE)
    ws.cell(row=r, column=9).fill = FILL_ACCENT
    ws.cell(row=r, column=9).number_format = NUM_EUR
    ws.cell(row=r, column=9).alignment = Alignment(horizontal="center", vertical="center")
    for col in [2, 4, 6, 8]:
        ws.cell(row=r, column=col).font = FONT_NUM
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")

    r += 2
    ws.cell(row=r, column=1, value="Benchmark").font = FONT_LABEL_MUT
    stats = [("Moy", f"=AVERAGE({rec_bench_list})"), ("Méd", f"=MEDIAN({rec_bench_list})"),
             ("Min", f"=MIN({rec_bench_list})"), ("Max", f"=MAX({rec_bench_list})")]
    for i, (label, formula) in enumerate(stats):
        lbl = ws.cell(row=r, column=2 + i * 2, value=label)
        lbl.font = Font(name="Arial", size=8, italic=True, color=MUTED)
        lbl.alignment = Alignment(horizontal="right")
        val = ws.cell(row=r, column=3 + i * 2, value=formula)
        val.font = FONT_NUM
        val.number_format = NUM_EUR
        val.alignment = Alignment(horizontal="left", indent=1)
    median_cell = f"E{r}"

    r += 2
    headers = ["#", "PP TTC", "PP HT", "PVM net", "Marge €", "% Marge", "vs Médiane", "vs Concentrate", "Flag"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_MUTED
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 20

    r += 1
    options_start = r
    for i, offset in enumerate(OFFSETS):
        row = r + i
        sign = "+" if offset >= 0 else "-"
        abs_off = abs(offset)
        expr = f"{pp_opt_cell}" if offset == 0 else f"{pp_opt_cell}{sign}{abs_off}"
        ws.cell(row=row, column=1, value=i + 1).font = Font(name="Arial", size=10, bold=True, color=MUTED)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=f"=CEILING({expr}+0.1,1)-0.1").number_format = NUM_EUR
        ws.cell(row=row, column=3, value=f"=B{row}/(1+{tva_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=4, value=f"=B{row}*(1-{remise_cell})").number_format = NUM_EUR
        ws.cell(row=row, column=5, value=f"=D{row}-{pr_cell}").number_format = NUM_EUR
        ws.cell(row=row, column=6, value=f"=IFERROR(E{row}/D{row},0)").number_format = NUM_PCT
        ws.cell(row=row, column=7, value=f"=IFERROR((B{row}-{median_cell})/{median_cell},0)").number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=8, value=f"=IFERROR((B{row}-{concentrate_retenu_cell})/{concentrate_retenu_cell},0)").number_format = NUM_PCT_SIGNED
        ws.cell(row=row, column=9, value=flag_formula_cascading(f"F{row}", f"B{row}", concentrate_retenu_cell, 0.05))
        for col in range(2, 9):
            ws.cell(row=row, column=col).font = FONT_NUM
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right", indent=1)
        if i % 2 == 1:
            for col in range(1, 10):
                c = ws.cell(row=row, column=col)
                if not c.fill.start_color or c.fill.start_color.value in (None, "00000000"):
                    c.fill = FILL_LIGHT
    options_end = r + 9

    r += 11
    ws.cell(row=r, column=1, value="🎯 Option sélectionnée").font = FONT_LABEL
    sel_cell = f"B{r}"
    add_dropdown(ws, sel_cell)
    ws.row_dimensions[r].height = 26
    ws.cell(row=r, column=3, value="→ Prix retenu").font = FONT_LABEL
    ws.cell(row=r, column=4, value=f"=INDEX($B${options_start}:$B${options_end},${sel_cell})")
    ws.cell(row=r, column=4).font = FONT_PRIX_BIG
    ws.cell(row=r, column=4).fill = FILL_AMBER
    ws.cell(row=r, column=4).number_format = NUM_EUR
    ws.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=5, value="Marge").font = FONT_LABEL
    ws.cell(row=r, column=6, value=f"=INDEX($F${options_start}:$F${options_end},${sel_cell})")
    ws.cell(row=r, column=6).font = FONT_MARGE_BIG
    ws.cell(row=r, column=6).fill = FILL_AMBER
    ws.cell(row=r, column=6).number_format = NUM_PCT
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")


# ═════════════════════════════════════════════════════════════════════════
def main():
    DST.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(SRC, DST)
    wb = openpyxl.load_workbook(DST)

    # Clean up old pricing tabs from v5 (removed previously)
    for old_name in ("Pricing Whey Concentrate", "Pricing Whey Isolate", "Pricing Whey Recovery", "Pricing Dynamique"):
        if old_name in wb.sheetnames:
            del wb[old_name]

    # Build new tabs
    build_variation_c(wb)
    build_variation_a(wb)

    wb.save(DST)
    print(f"✓ Generated {DST}")
    print(f"  Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
