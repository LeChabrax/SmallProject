"""
Build benchmark_whey_concurrent.xlsx from whey_catalog.json + discounts.json.

4 onglets :
 1. Benchmark détaillé — format du xlsx Vitavea (attributs en lignes, SKUs en colonnes).
    Les prix dérivés (€/kg, €/portion, €/g prot, prix effectif P50/P90) sont des FORMULES EXCEL
    qui référencent les cellules brutes (prix catalogue, format, servings, protéines/portion).
    Les remises P10/P50/P90/Max sont des VLOOKUP vers l'onglet "Discounts par marque".
    → modifier n'importe quel input (prix, format, remise, etc.) recalcule tout live dans le xlsx.
 2. Synthèse prix — même logique : inputs en colonnes E à H, dérivés en formules.
 3. Discounts par marque — source de vérité pour les remises (modifiable par l'utilisateur).
 4. Classement €/g protéine — SNAPSHOT figé au build : l'ordre du tri est calculé une fois
    puis écrit comme valeurs. Regénère le xlsx (python3 build_whey_xlsx.py) pour rafraîchir.
"""

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
CATALOG = HERE / "whey_catalog.json"
DISCOUNTS = HERE / "discounts.json"
OUT = HERE / "benchmark_whey_concurrent.xlsx"

BRAND_ORDER = [
    "impulse_nutrition",
    "nutripure",
    "nutrimuscle",
    "nutriandco",
    "aqeelab",
    "eric_favre",
    "eafit",
    "decathlon",
    "bulk",
    "myprotein",
    "prozis",
    "foodspring",
    "esn",
    "biotechusa",
]

BRAND_DISPLAY = {
    "impulse_nutrition": "IMPULSE",
    "nutripure": "NUTRIPURE",
    "nutrimuscle": "NUTRIMUSCLE",
    "nutriandco": "NUTRI&CO",
    "aqeelab": "AQEELAB",
    "eric_favre": "ERIC FAVRE",
    "eafit": "EAFIT",
    "decathlon": "DECATHLON",
    "bulk": "BULK",
    "myprotein": "MYPROTEIN",
    "prozis": "PROZIS",
    "foodspring": "FOODSPRING",
    "esn": "ESN",
    "biotechusa": "BIOTECHUSA",
}

# Name of the Discounts sheet as referenced in VLOOKUP formulas
DISCOUNTS_SHEET = "Discounts par marque"
# Columns of the Discounts sheet (1-indexed): A=Marque, B=1ère cmd, C=P10, D=P50, E=P90, F=Max, G=Prix eff %
DISCOUNTS_COL_P10 = 3
DISCOUNTS_COL_P50 = 4
DISCOUNTS_COL_P90 = 5
DISCOUNTS_COL_MAX = 6

# ── Styles
HEADER_BRAND_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_SECTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_SUBSECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
IMPULSE_HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
IMPULSE_BRAND_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
FORMULA_TINT_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD = Font(bold=True)
SECTION_FONT = Font(color="FFFFFF", bold=True, size=12)
FORMULA_FONT = Font(italic=True, color="1F3864")
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
CENTER = Alignment(horizontal="center", vertical="center")


# ──────────────────────────────────────────────────────────────
#  Helpers
# ──────────────────────────────────────────────────────────────
def load():
    catalog = json.loads(CATALOG.read_text())
    discounts = json.loads(DISCOUNTS.read_text()) if DISCOUNTS.exists() else {"discounts": {}}
    return catalog, discounts


def legacy_fallback(d, key):
    """Returns d[key] if set, else influencer_code_typical_pct (legacy brands)."""
    if not d:
        return None
    if d.get(key) is not None:
        return d.get(key)
    return d.get("influencer_code_typical_pct")


def _top_example_code(brand_discount):
    if not brand_discount:
        return None
    examples = brand_discount.get("influencer_code_examples") or []
    if not examples:
        return None
    top = max(examples, key=lambda e: e.get("pct") or 0)
    return f"{top.get('code', '?')} ({top.get('pct', '?')}%)"


def format_label(p):
    fmt = p.get("format_g")
    if fmt is None:
        return ""
    if fmt >= 1000:
        return f"{fmt/1000:g}kg" if fmt % 1000 else f"{fmt//1000}kg"
    return f"{fmt}g"


def type_label(p):
    t = (p.get("type") or "").replace("_", " ")
    if p.get("native_vs_fromagere") == "native":
        t += " (natif)"
    return t.strip()


def sort_products(products):
    type_rank = {
        "isolat_natif": 0,
        "isolat_fromagere": 1,
        "concentrat_natif": 2,
        "concentrat_fromagere": 3,
        "mix_concentrat_isolat": 4,
        "clear_hydrolysat": 5,
    }
    brand_rank = {b: i for i, b in enumerate(BRAND_ORDER)}
    return sorted(
        products,
        key=lambda p: (
            brand_rank.get(p["brand"], 99),
            type_rank.get(p.get("type") or "", 9),
            p.get("format_g") or 0,
        ),
    )


# ──────────────────────────────────────────────────────────────
#  Onglet 1 : Benchmark détaillé — pivot attributs × SKUs, formules live
# ──────────────────────────────────────────────────────────────
# kind values:
#   'header'          — header rows (brand name, product label)
#   'section'         — blue section separator row
#   'raw'             — input cell, value from getter(product)
#   'discount'        — VLOOKUP into Discounts sheet, brand name in row 1
#   'formula'         — Excel formula built from column letter + input row numbers

DETAILED_LAYOUT = [
    ('header_brand', 'MARQUE / PRODUIT'),
    ('header_ref', 'RÉFÉRENCE'),
    ('section', 'IDENTIFICATION'),
    ('raw', 'Type de whey', 'type_text', lambda p: type_label(p)),
    ('raw', 'Poids net (g)', 'format_g', lambda p: p.get('format_g')),
    ('raw', 'Arôme pivot', 'flavor', lambda p: p.get('flavor')),
    ('raw', 'Arômes disponibles (sample)', 'flavors', lambda p: ", ".join(p.get('flavors_available_sample') or [])),
    ('raw', 'Origine lait', 'origin_milk', lambda p: p.get('origin_milk')),
    ('raw', 'Certifications', 'certifications', lambda p: ", ".join(p.get('certifications') or [])),
    ('raw', 'Native / Fromagère', 'native_vs_fromagere', lambda p: p.get('native_vs_fromagere')),
    ('section', 'FORMULE & NUTRITION'),
    ('raw', '% protéines / 100g', 'protein_pct_100g', lambda p: p.get('protein_pct_100g')),
    ('raw', 'Protéines / portion (g)', 'protein_per_serving_g', lambda p: p.get('protein_per_serving_g')),
    ('raw', 'Taille portion (g)', 'serving_size_g', lambda p: p.get('serving_size_g')),
    ('raw', 'Nb portions / format', 'servings_per_pack', lambda p: p.get('servings_per_pack')),
    ('raw', 'BCAA / portion (g)', 'bcaa_per_serving_g', lambda p: p.get('bcaa_per_serving_g')),
    ('raw', 'Glucides / portion (g)', 'carbs_per_serving_g', lambda p: p.get('carbs_per_serving_g')),
    ('raw', 'Lipides / portion (g)', 'fat_per_serving_g', lambda p: p.get('fat_per_serving_g')),
    ('raw', 'Lactose', 'lactose_level', lambda p: p.get('lactose_level')),
    ('section', 'PRIX — 3 COUCHES (nominal / current / marché médian)'),
    ('raw', 'Prix nominal (site propre, barré)', 'price_nominal_ttc_eur', lambda p: p.get('price_nominal_ttc_eur')),
    ('raw', 'Prix actuel (site propre, affiché)', 'price_current_ttc_eur', lambda p: p.get('price_current_ttc_eur')),
    ('raw', 'Prix médian marché (idealo/ledenicheur)', 'price_median_market_ttc_eur', lambda p: p.get('price_median_market_ttc_eur')),
    ('formula', 'Écart nominal → actuel (%)',
     lambda col, r: f'=IFERROR(({col}{r["price_nominal_ttc_eur"]}-{col}{r["price_current_ttc_eur"]})/{col}{r["price_nominal_ttc_eur"]}*100,"")'),
    ('formula', 'Écart actuel → marché (%)',
     lambda col, r: f'=IFERROR(({col}{r["price_current_ttc_eur"]}-{col}{r["price_median_market_ttc_eur"]})/{col}{r["price_current_ttc_eur"]}*100,"")'),
    ('section', 'PRIX DÉRIVÉS (base = prix actuel site propre)'),
    ('formula', 'Prix actuel / kg (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}/({col}{r["format_g"]}/1000),"")'),
    ('formula', 'Prix actuel / portion (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}/{col}{r["servings_per_pack"]},"")'),
    ('formula', 'Prix actuel / g de protéine (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}/({col}{r["servings_per_pack"]}*{col}{r["protein_per_serving_g"]}),"")'),
    ('formula', 'Prix marché / g de protéine (€)',
     lambda col, r: f'=IFERROR({col}{r["price_median_market_ttc_eur"]}/({col}{r["servings_per_pack"]}*{col}{r["protein_per_serving_g"]}),"")'),
    ('raw', 'Prix abonnement (€ TTC)', 'price_subscription_ttc_eur', lambda p: p.get('price_subscription_ttc_eur')),
    ('raw', 'Remise abo (%)', 'subscription_discount_pct', lambda p: p.get('subscription_discount_pct')),
    ('section', 'PRIX EFFECTIF (avec code influenceur appliqué sur prix actuel)'),
    ('discount', 'Remise bas fourchette (%)', DISCOUNTS_COL_P10, 'remise_p10'),
    ('discount', 'Remise client standard (%)', DISCOUNTS_COL_P50, 'remise_p50'),
    ('discount', 'Remise client averti (%)', DISCOUNTS_COL_P90, 'remise_p90'),
    ('discount', 'Remise plafond observé (%)', DISCOUNTS_COL_MAX, 'remise_max'),
    ('formula', 'Prix client standard (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}*(1-{col}{r["remise_p50"]}/100),"")'),
    ('formula', 'Prix client averti (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}*(1-{col}{r["remise_p90"]}/100),"")'),
    ('formula', 'Prix plafond (€)',
     lambda col, r: f'=IFERROR({col}{r["price_current_ttc_eur"]}*(1-{col}{r["remise_max"]}/100),"")'),
    ('formula', 'Prix client standard / kg (€)',
     lambda col, r: f'=IFERROR({col}{r["price_effectif_p50"]}/({col}{r["format_g"]}/1000),"")'),
    ('formula', 'Prix client standard / g prot (€)',
     lambda col, r: f'=IFERROR({col}{r["price_effectif_p50"]}/({col}{r["servings_per_pack"]}*{col}{r["protein_per_serving_g"]}),"")'),
    ('formula', 'Prix client averti / g prot (€)',
     lambda col, r: f'=IFERROR({col}{r["price_effectif_p90"]}/({col}{r["servings_per_pack"]}*{col}{r["protein_per_serving_g"]}),"")'),
    ('section', 'DISTRIBUTION & SOURCE'),
    ('raw', 'Canaux de vente', 'channels', lambda p: ", ".join(p.get('channels') or [])),
    ('raw', 'Source URL (principale)', 'source_url', lambda p: p.get('source_url')),
    ('raw', 'Date snapshot', 'snapshot_date', lambda p: p.get('snapshot_date')),
    ('raw', 'Note gap catalogue/marché', 'price_gap_note', lambda p: p.get('price_gap_note')),
]

# Field name aliases so formula rows can reference previous formula rows
FORMULA_ROW_ALIASES = {
    'Prix client standard (€)': 'price_effectif_p50',
    'Prix client averti (€)': 'price_effectif_p90',
    'Prix plafond (€)': 'price_effectif_max',
}


DETAILED_ROW_MAP = {}  # populated by build_detailed_sheet — field_name → row_idx


def build_detailed_sheet(ws, products, discounts_data):
    global DETAILED_ROW_MAP
    products = sort_products(products)
    ws.column_dimensions["A"].width = 36
    # Assign row numbers for each field
    row_map = {}
    row_idx = 1
    for entry in DETAILED_LAYOUT:
        kind = entry[0]
        if kind == 'header_brand':
            row_map['__brand_row__'] = row_idx
        elif kind == 'header_ref':
            row_map['__ref_row__'] = row_idx
        elif kind == 'section':
            row_map[f'__section_{row_idx}__'] = row_idx
        elif kind == 'raw':
            _, _, field_name, _ = entry
            row_map[field_name] = row_idx
        elif kind == 'discount':
            _, _, _, alias = entry
            row_map[alias] = row_idx
        elif kind == 'formula':
            label = entry[1]
            alias = FORMULA_ROW_ALIASES.get(label)
            if alias:
                row_map[alias] = row_idx
        row_idx += 1

    DETAILED_ROW_MAP = row_map

    # Write content row-by-row
    row_idx = 1
    for entry in DETAILED_LAYOUT:
        kind = entry[0]

        if kind == 'header_brand':
            c = ws.cell(row=row_idx, column=1, value=entry[1])
            c.font = WHITE_FONT
            c.fill = HEADER_BRAND_FILL
            for col_idx, p in enumerate(products, start=2):
                cc = ws.cell(row=row_idx, column=col_idx, value=BRAND_DISPLAY.get(p["brand"], p["brand"]))
                cc.font = WHITE_FONT
                cc.alignment = CENTER
                cc.fill = IMPULSE_BRAND_FILL if p["brand"] == "impulse_nutrition" else HEADER_BRAND_FILL
                ws.column_dimensions[get_column_letter(col_idx)].width = 18

        elif kind == 'header_ref':
            c = ws.cell(row=row_idx, column=1, value=entry[1])
            c.font = BOLD
            c.fill = HEADER_SUBSECTION_FILL
            for col_idx, p in enumerate(products, start=2):
                label = f"{p['product_name']}\n({format_label(p)} {p.get('flavor', '')})"
                cc = ws.cell(row=row_idx, column=col_idx, value=label)
                cc.alignment = WRAP
                cc.font = BOLD
                if p["brand"] == "impulse_nutrition":
                    cc.fill = IMPULSE_HIGHLIGHT_FILL
            ws.row_dimensions[row_idx].height = 36

        elif kind == 'section':
            c = ws.cell(row=row_idx, column=1, value=entry[1])
            c.font = SECTION_FONT
            c.fill = HEADER_SECTION_FILL
            for col_idx in range(2, len(products) + 2):
                ws.cell(row=row_idx, column=col_idx).fill = HEADER_SECTION_FILL

        elif kind == 'raw':
            _, label, _, getter = entry
            c = ws.cell(row=row_idx, column=1, value=label)
            c.font = BOLD
            c.fill = HEADER_SUBSECTION_FILL
            for col_idx, p in enumerate(products, start=2):
                cc = ws.cell(row=row_idx, column=col_idx, value=getter(p))
                cc.alignment = WRAP
                if p["brand"] == "impulse_nutrition":
                    cc.fill = IMPULSE_HIGHLIGHT_FILL

        elif kind == 'discount':
            _, label, discount_col, _alias = entry
            c = ws.cell(row=row_idx, column=1, value=label)
            c.font = BOLD
            c.fill = HEADER_SUBSECTION_FILL
            brand_row = row_map['__brand_row__']
            for col_idx, p in enumerate(products, start=2):
                col_letter = get_column_letter(col_idx)
                formula = (
                    f'=IFERROR(VLOOKUP({col_letter}{brand_row},'
                    f"'{DISCOUNTS_SHEET}'!$A:$G,{discount_col},FALSE),0)"
                )
                cc = ws.cell(row=row_idx, column=col_idx, value=formula)
                cc.font = FORMULA_FONT
                cc.number_format = '0"%"'
                if p["brand"] == "impulse_nutrition":
                    cc.fill = IMPULSE_HIGHLIGHT_FILL
                else:
                    cc.fill = FORMULA_TINT_FILL

        elif kind == 'formula':
            _, label, formula_fn = entry
            c = ws.cell(row=row_idx, column=1, value=label)
            c.font = BOLD
            c.fill = HEADER_SUBSECTION_FILL
            for col_idx, p in enumerate(products, start=2):
                col_letter = get_column_letter(col_idx)
                formula = formula_fn(col_letter, row_map)
                cc = ws.cell(row=row_idx, column=col_idx, value=formula)
                cc.font = FORMULA_FONT
                if '€' in label or 'kg' in label or 'protéine' in label:
                    cc.number_format = '0.00" €"' if 'prot' not in label else '0.0000" €"'
                if p["brand"] == "impulse_nutrition":
                    cc.fill = IMPULSE_HIGHLIGHT_FILL
                else:
                    cc.fill = FORMULA_TINT_FILL

        row_idx += 1

    ws.freeze_panes = "B3"


# ──────────────────────────────────────────────────────────────
#  Onglet 2 : Synthèse prix — 1 SKU par ligne, formules live
# ──────────────────────────────────────────────────────────────
SYNTH_HEADERS = [
    "Marque", "Produit", "Type", "Natif/Fromagère",
    "Format (g)", "Nb portions", "Prot/portion (g)",
    # 3 price layers (inputs)
    "Prix nominal (€)", "Prix actuel (€)", "Prix marché (€)",
    # Gap columns
    "Gap nominal→actuel (%)", "Gap actuel→marché (%)",
    # Derived from price_current
    "Prix actuel /kg", "Prix actuel /g prot",
    "Prix abo (€)", "% abo",
    # Effective with discount (renommé : P50 = client standard, P90 = client averti)
    "Remise client standard (%)", "Remise client averti (%)",
    "Prix client standard (€)", "Prix client standard /g prot",
    "Prix client averti (€)", "Prix client averti /g prot",
    # Metadata
    "Source URL", "Gap note",
]
SYNTH_COL = {h: i + 1 for i, h in enumerate(SYNTH_HEADERS)}


SYNTH_ROW_MAP = {}  # populated by build_synthesis_sheet — product_id → synth_row


def build_synthesis_sheet(ws, products):
    global SYNTH_ROW_MAP
    SYNTH_ROW_MAP = {}
    for col_idx, h in enumerate(SYNTH_HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = WHITE_FONT
        c.fill = HEADER_SECTION_FILL
        c.alignment = CENTER
    products = sort_products(products)
    # Column letter shortcuts
    COL_MARQUE = get_column_letter(SYNTH_COL["Marque"])
    COL_FORMAT = get_column_letter(SYNTH_COL["Format (g)"])
    COL_SERVINGS = get_column_letter(SYNTH_COL["Nb portions"])
    COL_PROT = get_column_letter(SYNTH_COL["Prot/portion (g)"])
    COL_NOMINAL = get_column_letter(SYNTH_COL["Prix nominal (€)"])
    COL_CURRENT = get_column_letter(SYNTH_COL["Prix actuel (€)"])
    COL_MARKET = get_column_letter(SYNTH_COL["Prix marché (€)"])
    COL_REMISE_P50 = get_column_letter(SYNTH_COL["Remise client standard (%)"])
    COL_REMISE_P90 = get_column_letter(SYNTH_COL["Remise client averti (%)"])
    COL_EFF_P50 = get_column_letter(SYNTH_COL["Prix client standard (€)"])
    COL_EFF_P90 = get_column_letter(SYNTH_COL["Prix client averti (€)"])

    # Benchmark détaillé reference helper : for row r in Synthèse, the SKU is at column letter(r) in Benchmark
    # (both sheets use sort_products() so order matches — Synthèse row 2 == Benchmark col B == sku index 0)
    BENCH = "'Benchmark détaillé'"
    rm = DETAILED_ROW_MAP  # field_name -> row number in Benchmark détaillé

    for row_idx, p in enumerate(products, start=2):
        SYNTH_ROW_MAP[p["product_id"]] = row_idx
        bench_col = get_column_letter(row_idx)  # Benchmark détaillé column letter for this SKU

        # RAW inputs = formula references to Benchmark détaillé (single source of truth for raw data)
        ws.cell(row=row_idx, column=SYNTH_COL["Marque"],
                value=f"={BENCH}!{bench_col}{rm['__brand_row__']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Produit"], value=p.get("product_name"))
        ws.cell(row=row_idx, column=SYNTH_COL["Type"],
                value=f"={BENCH}!{bench_col}{rm['type_text']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Natif/Fromagère"],
                value=f"={BENCH}!{bench_col}{rm['native_vs_fromagere']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Format (g)"],
                value=f"={BENCH}!{bench_col}{rm['format_g']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Nb portions"],
                value=f"={BENCH}!{bench_col}{rm['servings_per_pack']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Prot/portion (g)"],
                value=f"={BENCH}!{bench_col}{rm['protein_per_serving_g']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Prix nominal (€)"],
                value=f"={BENCH}!{bench_col}{rm['price_nominal_ttc_eur']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Prix actuel (€)"],
                value=f"={BENCH}!{bench_col}{rm['price_current_ttc_eur']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Prix marché (€)"],
                value=f"={BENCH}!{bench_col}{rm['price_median_market_ttc_eur']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Prix abo (€)"],
                value=f"={BENCH}!{bench_col}{rm['price_subscription_ttc_eur']}")
        ws.cell(row=row_idx, column=SYNTH_COL["% abo"],
                value=f"={BENCH}!{bench_col}{rm['subscription_discount_pct']}")
        ws.cell(row=row_idx, column=SYNTH_COL["Source URL"], value=p.get("source_url"))
        ws.cell(row=row_idx, column=SYNTH_COL["Gap note"], value=p.get("price_gap_note"))

        # FORMULAS — reference same-row cells
        r = row_idx
        formulas = {
            "Gap nominal→actuel (%)":
                f'=IFERROR(({COL_NOMINAL}{r}-{COL_CURRENT}{r})/{COL_NOMINAL}{r}*100,"")',
            "Gap actuel→marché (%)":
                f'=IFERROR(({COL_CURRENT}{r}-{COL_MARKET}{r})/{COL_CURRENT}{r}*100,"")',
            "Prix actuel /kg":
                f'=IFERROR({COL_CURRENT}{r}/({COL_FORMAT}{r}/1000),"")',
            "Prix actuel /g prot":
                f'=IFERROR({COL_CURRENT}{r}/({COL_SERVINGS}{r}*{COL_PROT}{r}),"")',
            "Remise client standard (%)":
                f"=IFERROR(VLOOKUP({COL_MARQUE}{r},'{DISCOUNTS_SHEET}'!$A:$G,{DISCOUNTS_COL_P50},FALSE),0)",
            "Remise client averti (%)":
                f"=IFERROR(VLOOKUP({COL_MARQUE}{r},'{DISCOUNTS_SHEET}'!$A:$G,{DISCOUNTS_COL_P90},FALSE),0)",
            "Prix client standard (€)":
                f'=IFERROR({COL_CURRENT}{r}*(1-{COL_REMISE_P50}{r}/100),"")',
            "Prix client standard /g prot":
                f'=IFERROR({COL_EFF_P50}{r}/({COL_SERVINGS}{r}*{COL_PROT}{r}),"")',
            "Prix client averti (€)":
                f'=IFERROR({COL_CURRENT}{r}*(1-{COL_REMISE_P90}{r}/100),"")',
            "Prix client averti /g prot":
                f'=IFERROR({COL_EFF_P90}{r}/({COL_SERVINGS}{r}*{COL_PROT}{r}),"")',
        }
        for header, formula in formulas.items():
            cc = ws.cell(row=r, column=SYNTH_COL[header], value=formula)
            cc.font = FORMULA_FONT
            if p["brand"] != "impulse_nutrition":
                cc.fill = FORMULA_TINT_FILL

        if p["brand"] == "impulse_nutrition":
            for col_idx in range(1, len(SYNTH_HEADERS) + 1):
                ws.cell(row=r, column=col_idx).fill = IMPULSE_HIGHLIGHT_FILL

    widths = [14, 30, 22, 13, 10, 12, 13, 13, 13, 13, 14, 14, 13, 16, 13, 8, 12, 12, 14, 16, 14, 16, 50, 30]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "B2"


# ──────────────────────────────────────────────────────────────
#  Onglet 3 : Discounts par marque (source de vérité pour VLOOKUP)
# ──────────────────────────────────────────────────────────────
def build_discounts_sheet(ws, discounts_data):
    headers = [
        "Marque",                       # A  (used as VLOOKUP key)
        "1ère cmd (%)",                 # B
        "Remise bas fourchette (%)",    # C  (DISCOUNTS_COL_P10)
        "Remise client standard (%)",   # D  (DISCOUNTS_COL_P50)
        "Remise client averti (%)",     # E  (DISCOUNTS_COL_P90)
        "Remise plafond observé (%)",   # F  (DISCOUNTS_COL_MAX)
        "% prix réel vs catalogue",     # G
        "Fréquence promo",              # H
        "Codes (top 5)",                # I
        "N codes sourcés",              # J
        "Règles stacking",              # K
        "Note",                         # L
        "Sources",                      # M
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = WHITE_FONT
        c.fill = HEADER_SECTION_FILL
    discounts = discounts_data.get("discounts", {})
    row_idx = 2
    for brand in BRAND_ORDER:
        d = discounts.get(brand, {})
        examples = d.get("influencer_code_examples") or []
        top_examples = sorted(examples, key=lambda e: -(e.get("pct") or 0))[:5]
        example_str = "\n".join(
            f"{e.get('code', '?')} {e.get('pct', '?')}% ({e.get('influencer', '?')})" for e in top_examples
        ) if top_examples else ""
        sources = "\n".join(d.get("source_urls") or [])
        p10 = legacy_fallback(d, "influencer_code_pct_p10")
        p50 = legacy_fallback(d, "influencer_code_pct_p50")
        p90 = legacy_fallback(d, "influencer_code_pct_p90")
        pmax = legacy_fallback(d, "influencer_code_pct_max")
        eff_pct = d.get("effective_price_pct")
        if eff_pct is None and p50 is not None:
            eff_pct = 100 - p50
        values = [
            BRAND_DISPLAY.get(brand, brand),
            d.get("first_order_code_pct"),
            p10, p50, p90, pmax,
            eff_pct,
            d.get("permanent_promo_frequency"),
            example_str,
            len(examples),
            d.get("stacking_rules"),
            d.get("note"),
            sources,
        ]
        for col_idx, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=v)
            c.alignment = WRAP
        ws.row_dimensions[row_idx].height = 95
        row_idx += 1
    widths = [14, 10, 16, 16, 16, 16, 18, 16, 34, 10, 30, 40, 40]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "B2"


# ──────────────────────────────────────────────────────────────
#  Onglet 4 : Classement €/g protéine — SNAPSHOT (tri figé au build)
# ──────────────────────────────────────────────────────────────
def _price_per_g_protein(p, price_field="price_current_ttc_eur"):
    price = p.get(price_field)
    servings = p.get("servings_per_pack")
    prot = p.get("protein_per_serving_g")
    if price and servings and prot:
        return round(price / (servings * prot), 4)
    return None


def _effective_price(p, discount_pct, price_field="price_current_ttc_eur"):
    price = p.get(price_field)
    if not price:
        return None
    return round(price * (1 - discount_pct / 100), 2)


def _effective_price_per_g_protein(p, discount_pct):
    eff = _effective_price(p, discount_pct)
    servings = p.get("servings_per_pack")
    prot = p.get("protein_per_serving_g")
    if eff and servings and prot:
        return round(eff / (servings * prot), 4)
    return None


def _discount_pct(brand_discount, quantile="p50"):
    if not brand_discount:
        return 0
    key = f"influencer_code_pct_{quantile}"
    if brand_discount.get(key) is not None:
        return brand_discount[key]
    return brand_discount.get("influencer_code_typical_pct") or 0


def build_ranking_sheet(ws, products, discounts_data):
    note = (
        "Classement par 'Prix client standard / g protéine' croissant "
        "(#1 = meilleur rapport qualité/prix pour un client avec un code ambassadeur typique). "
        "Les valeurs chiffrées sont des formules live référençant l'onglet 'Synthèse prix'. "
        "Seul l'ordre de tri est figé au build. Si tu modifies un prix/remise dans Synthèse ou Discounts, les valeurs ici se mettent à jour. "
        "Pour réordonner le classement après une modif majeure, regénère le xlsx."
    )
    ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    ws.row_dimensions[1].height = 55

    headers = [
        "Rang", "Marque", "Produit", "Format",
        "Prix nominal", "Prix actuel", "Prix marché",
        "Prix actuel /g prot",
        "Prix client standard", "Prix client standard /g prot",
        "Prix client averti", "Prix client averti /g prot",
        "Remise standard (%)",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = WHITE_FONT
        c.fill = HEADER_SECTION_FILL

    discounts = discounts_data.get("discounts", {})
    ranked = []
    for p in products:
        brand_disc = discounts.get(p["brand"])
        eff_p50 = _discount_pct(brand_disc, "p50")
        eff_per_g_p50 = _effective_price_per_g_protein(p, eff_p50)
        if eff_per_g_p50 is None:
            continue
        ranked.append((eff_per_g_p50, p))
    ranked.sort(key=lambda x: x[0])

    # Reference helpers for Synthèse prix
    SYNTH = "'Synthèse prix'"
    COL_MARQUE = get_column_letter(SYNTH_COL["Marque"])
    COL_NOMINAL = get_column_letter(SYNTH_COL["Prix nominal (€)"])
    COL_CURRENT = get_column_letter(SYNTH_COL["Prix actuel (€)"])
    COL_MARKET = get_column_letter(SYNTH_COL["Prix marché (€)"])
    COL_ACTUEL_GP = get_column_letter(SYNTH_COL["Prix actuel /g prot"])
    COL_CS = get_column_letter(SYNTH_COL["Prix client standard (€)"])
    COL_CS_GP = get_column_letter(SYNTH_COL["Prix client standard /g prot"])
    COL_CA = get_column_letter(SYNTH_COL["Prix client averti (€)"])
    COL_CA_GP = get_column_letter(SYNTH_COL["Prix client averti /g prot"])
    COL_REMISE_CS = get_column_letter(SYNTH_COL["Remise client standard (%)"])

    for rank, (_eff_per_g_p50, p) in enumerate(ranked, start=1):
        row = rank + 2
        synth_row = SYNTH_ROW_MAP.get(p["product_id"])
        if synth_row is None:
            continue

        # Rank = static int, Produit + Format = static text (SKU-level, unchanging)
        # All other columns = formula references to Synthèse prix row
        ws.cell(row=row, column=1, value=rank)
        ws.cell(row=row, column=2, value=f"={SYNTH}!{COL_MARQUE}{synth_row}")
        ws.cell(row=row, column=3, value=p.get("product_name"))
        ws.cell(row=row, column=4, value=format_label(p))
        ws.cell(row=row, column=5, value=f"={SYNTH}!{COL_NOMINAL}{synth_row}")
        ws.cell(row=row, column=6, value=f"={SYNTH}!{COL_CURRENT}{synth_row}")
        ws.cell(row=row, column=7, value=f"={SYNTH}!{COL_MARKET}{synth_row}")
        ws.cell(row=row, column=8, value=f"={SYNTH}!{COL_ACTUEL_GP}{synth_row}")
        ws.cell(row=row, column=9, value=f"={SYNTH}!{COL_CS}{synth_row}")
        ws.cell(row=row, column=10, value=f"={SYNTH}!{COL_CS_GP}{synth_row}")
        ws.cell(row=row, column=11, value=f"={SYNTH}!{COL_CA}{synth_row}")
        ws.cell(row=row, column=12, value=f"={SYNTH}!{COL_CA_GP}{synth_row}")
        ws.cell(row=row, column=13, value=f"={SYNTH}!{COL_REMISE_CS}{synth_row}")

        if p["brand"] == "impulse_nutrition":
            for col_idx in range(1, 14):
                ws.cell(row=row, column=col_idx).fill = IMPULSE_HIGHLIGHT_FILL
        else:
            # Light tint to signal formula cells
            for col_idx in [5, 6, 7, 8, 9, 10, 11, 12, 13]:
                if not ws.cell(row=row, column=col_idx).fill.start_color or ws.cell(row=row, column=col_idx).fill.start_color.value == "00000000":
                    ws.cell(row=row, column=col_idx).fill = FORMULA_TINT_FILL
                    ws.cell(row=row, column=col_idx).font = FORMULA_FONT

    widths = [6, 14, 30, 10, 13, 13, 13, 18, 18, 22, 18, 22, 16]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A3"


# ──────────────────────────────────────────────────────────────
def main():
    catalog, discounts = load()
    products = catalog["products"]

    wb = openpyxl.Workbook()
    # Build Discounts sheet first so it exists when onglet 1 references it via VLOOKUP
    ws_disc = wb.active
    ws_disc.title = DISCOUNTS_SHEET
    build_discounts_sheet(ws_disc, discounts)

    ws1 = wb.create_sheet("Benchmark détaillé", 0)
    build_detailed_sheet(ws1, products, discounts)

    ws2 = wb.create_sheet("Synthèse prix", 1)
    build_synthesis_sheet(ws2, products)

    ws4 = wb.create_sheet("Classement €⁄g protéine")
    build_ranking_sheet(ws4, products, discounts)

    wb.save(OUT)
    print(f"✓ Généré {OUT} ({len(products)} SKUs, 4 onglets, formules live)")


if __name__ == "__main__":
    main()
